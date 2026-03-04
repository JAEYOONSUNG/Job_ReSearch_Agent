"""Export job data to Excel files with parsed, structured columns."""

from __future__ import annotations

import json
import logging
import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd

from src.config import (
    EXCEL_OUTPUT_DIR, EXCLUDE_KEYWORDS, EXCLUDE_TITLE_KEYWORDS,
    FACULTY_TITLE_KEYWORDS, GARBAGE_TITLE_PATTERNS,
)
from src.db import _AGGREGATOR_INSTITUTES, get_all_pis, get_jobs, get_recommended_pis
from src.matching.job_parser import parse_structured_description
from src.matching.scorer import is_company

logger = logging.getLogger(__name__)

PI_COLUMNS = [
    "PI Name",
    "Institute",
    "Country",
    "Tier",
    "h-index",
    "Citations",
    "Fields",
    "Connected Seeds",
    "Recommendation Score",
    "Lab URL",
    "Scholar URL",
]


# ---------------------------------------------------------------------------
# Text cleaning helpers
# ---------------------------------------------------------------------------

def _clean_text(text: str | None, max_len: int = 500) -> str:
    """Strip markdown, excessive whitespace, and truncate.

    Flattens all newlines into ``|`` section dividers for readable
    single-line display.  Adjacent short lines are merged; headers
    (``### Header`` or ``Header =====``) become ``| HEADER |``.
    """
    if not text:
        return ""
    # Remove markdown bold/italic
    text = re.sub(r"\*{1,3}(.+?)\*{1,3}", r"\1", text)
    # Remove markdown header markers: ### Header
    text = re.sub(r"^#{1,4}\s*", "", text, flags=re.MULTILINE)
    # Convert "Header ====+" or "====+" underline-style headers to section break
    text = re.sub(r"^(.+?)\s*={4,}\s*$", r"| \1 |", text, flags=re.MULTILINE)
    text = re.sub(r"^={4,}\s*$", "", text, flags=re.MULTILINE)
    # Convert "---+" horizontal rules to section break
    text = re.sub(r"^-{4,}\s*$", "|", text, flags=re.MULTILINE)
    # Remove markdown links but keep text
    text = re.sub(r"\[(.+?)\]\(.+?\)", r"\1", text)
    # Remove HTML tags
    text = re.sub(r"<[^>]+>", " ", text)
    # Unescape backslash-escaped characters (e.g. \- from Indeed)
    text = re.sub(r"\\([^\\])", r"\1", text)
    # Replace paragraph breaks (2+ newlines) with " | "
    text = re.sub(r"\s*\n\s*\n\s*", " | ", text)
    # Replace single newlines with space
    text = text.replace("\n", " ").replace("\r", " ")
    # Collapse all whitespace
    text = re.sub(r"\s+", " ", text).strip()
    # Clean up redundant pipe separators
    text = re.sub(r"\|\s*\|", "|", text)
    text = re.sub(r"^\s*\|\s*", "", text)
    text = re.sub(r"\s*\|\s*$", "", text)
    return text[:max_len]


_MAX_PAPER_COLS = 5  # number of individual paper columns


def _parse_papers(papers_json: str | None) -> list[dict]:
    """Parse JSON paper list into a list of paper dicts."""
    if not papers_json:
        return []
    try:
        papers = json.loads(papers_json)
    except (json.JSONDecodeError, TypeError):
        return []
    return papers if isinstance(papers, list) else []


def _papers_to_columns(prefix: str, papers: list[dict]) -> dict[str, str]:
    """Convert a list of paper dicts to individual column values.

    Returns e.g. {"Recent Paper 1": "(2024) Title [10 cites]", ...}.
    The display text is used as a placeholder; actual hyperlinks are written
    by _write_paper_links() after the DataFrame is exported.
    """
    result: dict[str, str] = {}
    for i in range(_MAX_PAPER_COLS):
        col_name = f"{prefix} {i + 1}"
        if i < len(papers):
            p = papers[i]
            year = p.get("year") or "?"
            title = (p.get("title") or "Untitled")[:80]
            cites = p.get("citation_count", 0)
            result[col_name] = f"({year}) {title} [{cites} cites]"
        else:
            result[col_name] = ""
    return result


def _clean_list(text: str | None, max_len: int = 400) -> str:
    """Clean a list-style section (requirements, etc).

    Flattens bullets into semicolon-separated items for compact display.
    """
    if not text:
        return ""
    # Remove markdown bold/italic
    text = re.sub(r"\*{1,3}(.+?)\*{1,3}", r"\1", text)
    # Remove markdown headers
    text = re.sub(r"^#{1,4}\s*", "", text, flags=re.MULTILINE)
    # Remove underline-style headers
    text = re.sub(r"={4,}", "", text)
    text = re.sub(r"-{4,}", "", text)
    # Remove HTML tags
    text = re.sub(r"<[^>]+>", " ", text)
    # Unescape backslash-escaped characters
    text = re.sub(r"\\([^\\])", r"\1", text)
    # Normalise bullet markers to "; "
    text = re.sub(r"\n\s*[-•*]\s*", "; ", text)
    text = re.sub(r"\n\s*\d+[.)]\s*", "; ", text)
    # Replace remaining newlines with spaces
    text = text.replace("\n", " ").replace("\r", " ")
    # Collapse whitespace
    text = re.sub(r"\s+", " ", text).strip()
    # Clean up leading/trailing semicolons
    text = text.strip("; ").strip()
    return text[:max_len]


# ---------------------------------------------------------------------------
# Condition parsing — split into sub-fields
# ---------------------------------------------------------------------------

def _format_salary_number(s: str) -> str:
    """Add thousand-separator commas to bare salary numbers.

    ``$56484`` → ``$56,484``;  ``€45000-€55000`` → ``€45,000-€55,000``
    ``3800`` → ``3,800``;  ``42.185,28`` → ``42,185``
    Leaves already-formatted strings (``$56,484``) and non-numeric
    salary scales (``TV-L E13``) unchanged.
    """
    # First, normalise EU thousands separator (42.185,28 → 42185.28)
    s = re.sub(r"(\d{1,3})\.(\d{3}),(\d{2})\b", lambda m: f"{m.group(1)}{m.group(2)}", s)
    # Also handle "42.185" standalone (EU thousands, no decimals)
    s = re.sub(r"\b(\d{1,3})\.(\d{3})\b(?![\d.])", lambda m: f"{m.group(1)}{m.group(2)}", s)

    def _add_commas(m: re.Match) -> str:
        return f"{int(m.group()):,}"

    # Currency-prefixed: 4+ digits after $, €, £
    s = re.sub(r"(?<=[\$€£])\d{4,}", _add_commas, s)
    # Bare numbers: 4+ digits (not preceded/followed by digit or dot)
    s = re.sub(r"(?<![.\d])\b(\d{4,})\b(?![\d.])", lambda m: f"{int(m.group(1)):,}", s)
    # Strip decimal places from numbers (e.g. $60,000.00 → $60,000, 36.0 → 36)
    s = re.sub(r"(\d)\.0{1,2}\b", r"\1", s)          # .0 or .00
    s = re.sub(r"(\d)\.\d{1,2}(?=\s*[€$£%])", r"\1", s)  # 42,392.88€ → 42,392€
    s = re.sub(r"(\d)\.\d{1,2}(?=\s*/)", r"\1", s)   # $60,000.00/yr → $60,000/yr
    # Large comma-formatted numbers: strip remaining decimals (€ 70,167.44 etc)
    s = re.sub(r"(\d{1,3}(?:,\d{3})+)\.\d{1,2}\b", r"\1", s)
    return s


def _parse_salary(conditions: str, description: str) -> str:
    """Extract salary info from conditions or description."""
    blob = f"{conditions} {description}"
    patterns = [
        r"(?:salary|stipend|compensation|remuneration)\s*(?:of|is|:)?\s*([\$€£]?\s*\d[\d,]*(?:\.\d+)?(?:\s*[-–]\s*[\$€£]?\s*\d[\d,]*(?:\.\d+)?)?)\s*(?:k|K|per\s+(?:year|annum|month)|/\s*(?:yr|year|annum|month(?:ly)?)|euros?(?:/month)?)?",
        r"([\$€£]\s*\d[\d,]*(?:\.\d+)?(?:\s*[-–]\s*[\$€£]?\s*\d[\d,]*(?:\.\d+)?)?)\s*(?:k|K)?\s*(?:per\s+(?:year|annum)|/\s*(?:yr|year|annum|yearly))",
        r"(TV-?L\s*E?\s*1[3-5](?:\s*/\s*E?\s*1[3-5])?)",
        r"(NIH\s+(?:scale|salary))",
        r"(Grade\s+\d+|Band\s+\d+|Scale\s+\d+)",
        r"Salary:\s*([\$€£]?\d[\d,.]*(?:\s*[-–]\s*[\$€£]?\d[\d,.]*)?(?:/\w+)?)",
        r"(\d[\d.]*,\d{2}\s*€)",  # EU format: "42.185,28 €"
    ]
    for p in patterns:
        m = re.search(p, blob, re.IGNORECASE)
        if m:
            raw = m.group(1).strip() if m.lastindex else m.group(0).strip()
            return _format_salary_number(raw)
    return ""


def _parse_duration(conditions: str, description: str) -> str:
    """Extract duration/contract length."""
    blob = f"{conditions} {description}"
    patterns = [
        r"(\d+(?:\s*[-–]\s*\d+)?)\s*(?:year|yr)s?\s*(?:position|appointment|contract|renewable)?",
        r"(\d+(?:\s*[-–]\s*\d+)?)\s*months?\s*(?:position|appointment|contract)?",
        r"(?:up to|minimum|at least|initially)\s+(\d+)\s+(?:year|yr|month)s?",
        r"(?:duration|length|term)\s*(?:of|:)\s*(\d+\s*(?:year|yr|month)s?)",
        r"Contract:\s*(\w[\w\s,-]*)",
    ]
    for p in patterns:
        m = re.search(p, blob, re.IGNORECASE)
        if m:
            return m.group(0).strip()[:60]
    return ""


def _parse_contract_type(conditions: str, description: str) -> str:
    """Extract contract type (full-time, fixed-term, etc)."""
    blob = f"{conditions} {description}".lower()
    types = []
    if "full-time" in blob or "full time" in blob or "fulltime" in blob:
        types.append("Full-time")
    elif "part-time" in blob or "part time" in blob:
        types.append("Part-time")
    if "fixed-term" in blob or "fixed term" in blob or "temporary" in blob:
        types.append("Fixed-term")
    elif "permanent" in blob or "tenure" in blob:
        types.append("Permanent")
    return ", ".join(types) if types else ""


def _parse_start_date(conditions: str, description: str) -> str:
    """Extract start date."""
    blob = f"{conditions} {description}"
    patterns = [
        r"(?:start(?:ing)?\s*(?:date)?|commencement|begin(?:ning)?)\s*[:=]?\s*(\w+\s+\d{1,2},?\s+\d{4})",
        r"(?:start(?:ing)?|available|begin)\s*(?:date)?\s*[:=]?\s*((?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4})",
        r"(?:start(?:ing)?)\s*[:=]?\s*(\d{4}-\d{2}-\d{2})",
        r"(?:as soon as possible|immediately|ASAP)",
    ]
    for p in patterns:
        m = re.search(p, blob, re.IGNORECASE)
        if m:
            return m.group(1).strip() if m.lastindex else m.group(0).strip()
    return ""


# ---------------------------------------------------------------------------
# Requirements parsing — split into sub-fields
# ---------------------------------------------------------------------------

def _parse_degree(requirements: str, description: str) -> str:
    """Extract required degree."""
    blob = f"{requirements} {description}"
    patterns = [
        r"(Ph\.?D\.?\s+in\s+[\w\s,/&]+?)(?:\.|;|\n|$)",
        r"((?:Doctoral|PhD|Ph\.D)\s+(?:degree\s+)?in\s+[\w\s,/&]+?)(?:\.|;|\n|$)",
        r"((?:Master'?s?|M\.?S\.?|M\.?Sc\.?)\s+(?:degree\s+)?in\s+[\w\s,/&]+?)(?:\.|;|\n|$)",
    ]
    for p in patterns:
        m = re.search(p, blob, re.IGNORECASE)
        if m:
            return m.group(1).strip()[:120]
    return ""


def _parse_skills(requirements: str, description: str) -> str:
    """Extract required skills/techniques."""
    blob = f"{requirements} {description}".lower()
    skills = []
    skill_keywords = [
        "CRISPR", "Cas9", "Cas12", "flow cytometry", "FACS",
        "mass spectrometry", "NGS", "PCR", "qPCR", "RT-PCR",
        "Western blot", "microscopy", "confocal", "cryo-EM",
        "Python", "R programming", "bioinformatics", "machine learning",
        "cell culture", "cloning", "protein purification",
        "HPLC", "NMR", "X-ray crystallography",
        "mouse model", "animal model", "in vivo",
        "single-cell", "RNA-seq", "ChIP-seq", "ATAC-seq",
        "electrophysiology", "patch clamp", "optogenetics",
        "fermentation", "bioprocess", "metabolic flux",
        "directed evolution", "high-throughput screening",
        "gene editing", "genome engineering",
    ]
    for skill in skill_keywords:
        if skill.lower() in blob:
            skills.append(skill)
    return ", ".join(skills[:10]) if skills else ""


# ---------------------------------------------------------------------------
# Main row builder — parsed, pivoted columns
# ---------------------------------------------------------------------------

JOB_COLUMNS = [
    "Tier",
    "Title",
    "PI Name",
    "Institute",
    "Department",
    "Country",
    "Region",
    # Research
    "Field",
    "Keywords",
    # Structured sections (parsed from LinkedIn/Indeed descriptions)
    "Position Summary",
    "Responsibilities",
    "Preferred Qualifications",
    # Requirements (parsed)
    "Degree Required",
    "Skills/Techniques",
    "Requirements (Full)",
    # Conditions (parsed)
    "Salary",
    "Duration",
    "Contract Type",
    "Start Date",
    "Conditions (Full)",
    # Dates
    "Posted Date",
    "Deadline",
    "Application Materials",
    # Description
    "Description",
    # PI Papers — individual columns
    *[f"Recent Paper {i+1}" for i in range(_MAX_PAPER_COLS)],
    *[f"Top Cited Paper {i+1}" for i in range(_MAX_PAPER_COLS)],
    # Contact & info
    "Contact Email",
    "Info URL 1",
    "Info URL 2",
    "Info URL 3",
    # Links
    "Job URL",
    "Job URL 2",
    "Lab URL",
    "Scholar URL",
    "Dept URL",
    # Meta
    "Match Score",
    "Source",
    "Status",
]


_MAX_INFO_URLS = 3  # number of Info URL columns


def _info_urls_to_columns(info_urls_json: str | None) -> dict[str, str]:
    """Convert info_urls JSON to individual column values."""
    urls: list[str] = []
    if info_urls_json:
        try:
            urls = json.loads(info_urls_json)
        except (json.JSONDecodeError, TypeError):
            pass
    result: dict[str, str] = {}
    for i in range(_MAX_INFO_URLS):
        col = f"Info URL {i + 1}"
        result[col] = urls[i] if i < len(urls) else ""
    return result


def _format_tier(tier, institute: str = "") -> str:
    """Format tier for display: T1-T4 for institutions, 'Company' for companies, '' for unranked."""
    if institute and is_company(institute):
        return "Company"
    if not tier or tier >= 5:
        return ""
    return f"T{tier}"


def _job_to_row(job: dict) -> dict:
    tier = job.get("tier")
    institute = job.get("institute") or ""
    desc = job.get("description") or ""
    req = job.get("requirements") or ""
    cond = job.get("conditions") or ""

    # Parse structured sections from description (LinkedIn/Indeed format)
    sections = parse_structured_description(desc)

    # Use parsed requirements/conditions as fallback for empty fields
    if not req and sections.get("requirements"):
        req = sections["requirements"]
    if not cond and sections.get("conditions"):
        cond = sections["conditions"]

    return {
        "Title": (job.get("title") or "").strip(),
        "PI Name": job.get("pi_name") or "",
        "Institute": institute,
        "Department": job.get("department") or "",
        "Tier": _format_tier(tier, institute),
        "Country": job.get("country") or "",
        "Region": job.get("region") or "",
        # Research
        "Field": job.get("field") or "",
        "Keywords": job.get("keywords") or "",
        # Structured sections
        "Position Summary": _clean_text(sections.get("summary"), 2000),
        "Responsibilities": _clean_list(sections.get("responsibilities"), 2000),
        "Preferred Qualifications": _clean_list(sections.get("preferred"), 2000),
        # Requirements parsed
        "Degree Required": _parse_degree(req, desc),
        "Skills/Techniques": _parse_skills(req, desc),
        "Requirements (Full)": _clean_list(req, 3000),
        # Conditions parsed
        "Salary": _parse_salary(cond, desc),
        "Duration": _parse_duration(cond, desc),
        "Contract Type": _parse_contract_type(cond, desc),
        "Start Date": _parse_start_date(cond, desc),
        "Conditions (Full)": _format_salary_number(_clean_list(cond, 3000)),
        # Dates
        "Posted Date": job.get("posted_date") or "",
        "Deadline": job.get("deadline") or "",
        "Application Materials": _clean_list(job.get("application_materials"), 1000),
        # Description
        "Description": _clean_text(desc, 15000),
        # PI Papers — individual columns (title text; hyperlinks added in _write_paper_links)
        **_papers_to_columns("Recent Paper", _parse_papers(job.get("recent_papers"))),
        **_papers_to_columns("Top Cited Paper", _parse_papers(job.get("top_cited_papers"))),
        # Contact & info
        "Contact Email": job.get("contact_email") or "",
        **_info_urls_to_columns(job.get("info_urls")),
        # Links
        "Job URL": job.get("url") or "",
        "Job URL 2": job.get("alt_url") or "",
        "Lab URL": job.get("lab_url") or "",
        "Scholar URL": job.get("scholar_url") or "",
        "Dept URL": job.get("dept_url") or "",
        # Meta
        "Match Score": job.get("match_score", 0),
        "Source": job.get("source") or "",
        "Status": job.get("status") or "",
    }


def _pi_to_row(pi: dict) -> dict:
    institute = pi.get("institute", "")
    return {
        "PI Name": pi.get("name", ""),
        "Institute": institute,
        "Country": pi.get("country", ""),
        "Tier": _format_tier(pi.get("tier"), institute),
        "h-index": pi.get("h_index", ""),
        "Citations": pi.get("citations", ""),
        "Fields": pi.get("fields", ""),
        "Connected Seeds": pi.get("connected_seeds", ""),
        "Recommendation Score": pi.get("recommendation_score", 0),
        "Lab URL": pi.get("lab_url", ""),
        "Scholar URL": pi.get("scholar_url", ""),
    }


def _write_paper_links(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
    jobs: list[dict],
) -> None:
    """Overwrite paper cells with clickable hyperlinks pointing to S2/DOI URLs."""
    worksheet = writer.sheets[sheet_name]
    workbook = writer.book
    link_fmt = workbook.add_format({
        "font_color": "#0563C1",
        "underline": True,
        "valign": "vcenter",
        "text_wrap": True,
        "font_size": 9,
    })

    # Build column index lookup for paper columns
    paper_col_indices: dict[str, int] = {}
    for col_name in df.columns:
        if col_name.startswith("Recent Paper ") or col_name.startswith("Top Cited Paper "):
            paper_col_indices[col_name] = df.columns.get_loc(col_name)

    for row_idx, job in enumerate(jobs):
        excel_row = row_idx + 1  # +1 for header row

        # Recent papers
        recent = _parse_papers(job.get("recent_papers"))
        for i in range(min(_MAX_PAPER_COLS, len(recent))):
            col_name = f"Recent Paper {i + 1}"
            col_idx = paper_col_indices.get(col_name)
            if col_idx is None:
                continue
            p = recent[i]
            url = p.get("url", "")
            year = p.get("year") or "?"
            title = (p.get("title") or "Untitled")[:80]
            cites = p.get("citation_count", 0)
            display = f"({year}) {title} [{cites} cites]"
            if url:
                worksheet.write_url(excel_row, col_idx, url, link_fmt, display)
            else:
                worksheet.write(excel_row, col_idx, display)

        # Top cited papers
        top_cited = _parse_papers(job.get("top_cited_papers"))
        for i in range(min(_MAX_PAPER_COLS, len(top_cited))):
            col_name = f"Top Cited Paper {i + 1}"
            col_idx = paper_col_indices.get(col_name)
            if col_idx is None:
                continue
            p = top_cited[i]
            url = p.get("url", "")
            year = p.get("year") or "?"
            title = (p.get("title") or "Untitled")[:80]
            cites = p.get("citation_count", 0)
            display = f"({year}) {title} [{cites} cites]"
            if url:
                worksheet.write_url(excel_row, col_idx, url, link_fmt, display)
            else:
                worksheet.write(excel_row, col_idx, display)


def _style_worksheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """Apply formatting to a worksheet with auto-filters and conditional formatting."""
    worksheet = writer.sheets[sheet_name]
    workbook = writer.book

    # Header format
    header_fmt = workbook.add_format({
        "bold": True,
        "bg_color": "#16213e",
        "font_color": "white",
        "valign": "vcenter",
        "border": 1,
    })

    # Write headers with format
    for i, col in enumerate(df.columns):
        worksheet.write(0, i, col, header_fmt)

    # Column widths (tuned per column type)
    width_map = {
        "Title": 65, "PI Name": 18, "Institute": 25, "Department": 20,
        "Tier": 5, "Country": 12, "Region": 6,
        "Field": 20, "Keywords": 35,
        "Position Summary": 45, "Responsibilities": 45, "Preferred Qualifications": 40,
        "Degree Required": 30, "Skills/Techniques": 30, "Requirements (Full)": 40,
        "Salary": 18, "Duration": 18, "Contract Type": 12, "Start Date": 14,
        "Conditions (Full)": 30,
        "Posted Date": 11, "Deadline": 11,
        "Description": 50,
        **{f"Recent Paper {i+1}": 35 for i in range(_MAX_PAPER_COLS)},
        **{f"Top Cited Paper {i+1}": 35 for i in range(_MAX_PAPER_COLS)},
        "Contact Email": 25,
        **{f"Info URL {i+1}": 20 for i in range(_MAX_INFO_URLS)},
        "Job URL": 15, "Job URL 2": 15, "Lab URL": 15, "Scholar URL": 15, "Dept URL": 15,
        "Match Score": 8, "Source": 12, "Status": 7,
    }

    url_fmt = workbook.add_format({"font_color": "blue", "underline": True, "valign": "vcenter"})
    default_fmt = workbook.add_format({"valign": "vcenter"})

    for i, col in enumerate(df.columns):
        width = width_map.get(col, 15)
        if col in ("Job URL", "Job URL 2", "Lab URL", "Scholar URL", "Dept URL",
                   "Info URL 1", "Info URL 2", "Info URL 3"):
            worksheet.set_column(i, i, width, url_fmt)
        else:
            worksheet.set_column(i, i, width, default_fmt)

    # Compact row height
    worksheet.set_default_row(20)

    # Auto-filter on all columns
    if len(df) > 0:
        worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)

    # Conditional formatting: highlight Tier column
    if "Tier" in df.columns:
        tier_col = df.columns.get_loc("Tier")
        tier_range = f"{chr(65 + tier_col)}2:{chr(65 + tier_col)}{len(df) + 1}"
        # T1 = gold, T2 = blue, T3 = green, T4 = light gray
        worksheet.conditional_format(
            tier_range,
            {"type": "text", "criteria": "containing", "value": "T1",
             "format": workbook.add_format({"bg_color": "#FFF3CD", "font_color": "#856404", "bold": True})},
        )
        worksheet.conditional_format(
            tier_range,
            {"type": "text", "criteria": "containing", "value": "T2",
             "format": workbook.add_format({"bg_color": "#CCE5FF", "font_color": "#004085"})},
        )
        worksheet.conditional_format(
            tier_range,
            {"type": "text", "criteria": "containing", "value": "T3",
             "format": workbook.add_format({"bg_color": "#D4EDDA", "font_color": "#155724"})},
        )
        worksheet.conditional_format(
            tier_range,
            {"type": "text", "criteria": "containing", "value": "T4",
             "format": workbook.add_format({"bg_color": "#E2E3E5", "font_color": "#383D41"})},
        )
        worksheet.conditional_format(
            tier_range,
            {"type": "text", "criteria": "containing", "value": "Company",
             "format": workbook.add_format({"bg_color": "#E8DAEF", "font_color": "#6C3483", "bold": True})},
        )

    # Conditional formatting: color scale on Match Score
    if "Match Score" in df.columns:
        score_col = df.columns.get_loc("Match Score")
        worksheet.conditional_format(
            1, score_col, len(df), score_col,
            {"type": "3_color_scale",
             "min_color": "#F8D7DA", "mid_color": "#FFF3CD", "max_color": "#D4EDDA"},
        )

    # Conditional formatting: color scale on Recommendation Score (PI sheet)
    if "Recommendation Score" in df.columns:
        rec_col = df.columns.get_loc("Recommendation Score")
        worksheet.conditional_format(
            1, rec_col, len(df), rec_col,
            {"type": "3_color_scale",
             "min_color": "#F8D7DA", "mid_color": "#FFF3CD", "max_color": "#D4EDDA"},
        )

    # Freeze top row for easy scrolling
    worksheet.freeze_panes(1, 0)


def _deadline_bg_color(deadline_str: str) -> tuple[str, int] | tuple[None, int]:
    """Return (hex background color, days_remaining) based on deadline urgency.

    Smooth gradient: dark red #CC0000 (0 days) → light green #D5F5E3 (90+ days).
    Expired: gray #AAAAAA. Empty/unparseable: (None, 999).
    """
    if not deadline_str:
        return None, 999
    try:
        dl = date.fromisoformat(deadline_str)
    except (ValueError, TypeError):
        return None, 999

    days_remaining = (dl - date.today()).days
    if days_remaining < 0:
        return "#AAAAAA", days_remaining

    # Clamp to 0–90 range for interpolation
    t = min(days_remaining, 90) / 90.0
    # Dark red (0.80, 0.00, 0.00) → Light green (0.84, 0.96, 0.89)
    r = int((0.80 * (1 - t) + 0.84 * t) * 255)
    g = int((0.00 * (1 - t) + 0.96 * t) * 255)
    b = int((0.00 * (1 - t) + 0.89 * t) * 255)
    return f"#{r:02X}{g:02X}{b:02X}", days_remaining


def _style_deadline_cells(
    writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame
) -> None:
    """Apply urgency-based background colors to Deadline column cells."""
    if "Deadline" not in df.columns:
        return
    worksheet = writer.sheets[sheet_name]
    workbook = writer.book
    col_idx = df.columns.get_loc("Deadline")

    for row_idx in range(len(df)):
        val = df.iloc[row_idx, col_idx]
        deadline_str = str(val) if pd.notna(val) else ""
        color, days_remaining = _deadline_bg_color(deadline_str)
        if not color:
            continue

        fmt_props: dict = {
            "bg_color": color,
            "valign": "vcenter",
            "font_size": 9,
        }
        if days_remaining < 0:
            fmt_props["font_color"] = "#666666"
        elif days_remaining <= 7:
            fmt_props["bold"] = True
            fmt_props["font_color"] = "white"
        elif days_remaining <= 30:
            fmt_props["font_color"] = "white"

        cell_fmt = workbook.add_format(fmt_props)
        worksheet.write(row_idx + 1, col_idx, deadline_str, cell_fmt)


_CITIZENSHIP_PATTERNS = re.compile(
    r"U\.?S\.?\s*citizen|"
    r"citizenship\s+(?:is\s+)?require|"
    r"must\s+be\s+a?\s*(?:U\.?S\.?|American)\s*citizen|"
    r"U\.?S\.?\s*persons?\s+only|"
    r"(?:require|need)s?\s+(?:U\.?S\.?\s*)?(?:security\s+)?clearance|"
    r"\bITAR\b|"
    r"\bEAR\b.*(?:restrict|require|compliance)|"
    r"(?:only\s+)?(?:U\.?S\.?\s*)?(?:citizens?\s+(?:and|or)\s+)?permanent\s+residents?\s+(?:only|eligible|may\s+apply)",
    re.IGNORECASE,
)


def _has_citizenship_restriction(job: dict) -> bool:
    """Return True if the job requires specific citizenship/visa status."""
    blob = " ".join(
        (job.get(k) or "") for k in ("description", "conditions", "requirements")
    )
    return bool(_CITIZENSHIP_PATTERNS.search(blob))


def _style_citizenship_cells(
    writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame, jobs: list[dict],
) -> None:
    """Apply red text to the Institute cell when citizenship restrictions are detected."""
    if "Institute" not in df.columns:
        return
    worksheet = writer.sheets[sheet_name]
    workbook = writer.book
    col_idx = df.columns.get_loc("Institute")

    red_fmt = workbook.add_format({
        "font_color": "#CC0000",
        "bold": True,
        "valign": "vcenter",
    })

    for row_idx, job in enumerate(jobs):
        if _has_citizenship_restriction(job):
            val = df.iloc[row_idx, col_idx]
            institute_str = str(val) if pd.notna(val) else ""
            worksheet.write(row_idx + 1, col_idx, institute_str, red_fmt)


def _write_summary_dashboard(
    writer: pd.ExcelWriter,
    all_jobs: list[dict],
    rec_pis: list[dict],
) -> None:
    """Write a premium Summary Dashboard sheet with KPI cards, charts, and tables."""
    from collections import Counter

    workbook = writer.book
    ws = workbook.add_worksheet("Summary Dashboard")
    writer.sheets["Summary Dashboard"] = ws

    # ── Color palette ─────────────────────────────────────────────────
    NAVY = "#0F1B2D"
    ACCENT = "#3B82F6"     # bright blue
    ACCENT_DARK = "#1E40AF"
    GOLD = "#F59E0B"
    EMERALD = "#10B981"
    ROSE = "#F43F5E"
    SLATE = "#64748B"
    LIGHT_BG = "#F8FAFC"
    CARD_BG = "#FFFFFF"
    SUBTLE_BORDER = "#E2E8F0"
    TEXT_PRIMARY = "#1E293B"
    TEXT_SECONDARY = "#475569"
    TEXT_MUTED = "#94A3B8"

    # ── Reusable formats ──────────────────────────────────────────────
    bg_fmt = workbook.add_format({"bg_color": LIGHT_BG})

    header_fmt = workbook.add_format({
        "bold": True, "font_size": 22, "font_color": NAVY,
        "font_name": "Calibri", "bg_color": LIGHT_BG,
        "bottom": 0,
    })
    subtitle_fmt = workbook.add_format({
        "font_size": 10, "font_color": TEXT_MUTED,
        "font_name": "Calibri", "bg_color": LIGHT_BG,
    })
    section_title_fmt = workbook.add_format({
        "bold": True, "font_size": 13, "font_color": NAVY,
        "font_name": "Calibri", "bg_color": LIGHT_BG,
        "bottom": 2, "bottom_color": ACCENT,
    })

    # KPI card formats
    def _make_kpi_card(color: str):
        """Return (top_border, big_number, label, bg) formats for a KPI card."""
        top = workbook.add_format({
            "top": 5, "top_color": color,
            "bg_color": CARD_BG, "font_size": 1,
            "left": 1, "right": 1, "left_color": SUBTLE_BORDER, "right_color": SUBTLE_BORDER,
        })
        big = workbook.add_format({
            "bold": True, "font_size": 28, "font_color": color,
            "font_name": "Calibri", "align": "center", "valign": "vcenter",
            "bg_color": CARD_BG,
            "left": 1, "right": 1, "left_color": SUBTLE_BORDER, "right_color": SUBTLE_BORDER,
        })
        lbl = workbook.add_format({
            "font_size": 9, "font_color": TEXT_SECONDARY,
            "font_name": "Calibri", "align": "center", "valign": "top",
            "bg_color": CARD_BG,
            "left": 1, "right": 1, "bottom": 1,
            "left_color": SUBTLE_BORDER, "right_color": SUBTLE_BORDER,
            "bottom_color": SUBTLE_BORDER,
        })
        return top, big, lbl

    kpi_blue = _make_kpi_card(ACCENT)
    kpi_gold = _make_kpi_card(GOLD)
    kpi_emerald = _make_kpi_card(EMERALD)
    kpi_rose = _make_kpi_card(ROSE)

    # Table formats
    tbl_header_fmt = workbook.add_format({
        "bold": True, "font_size": 10, "font_color": "#FFFFFF",
        "bg_color": NAVY, "font_name": "Calibri",
        "align": "center", "valign": "vcenter",
        "top": 1, "bottom": 1, "left": 1, "right": 1,
        "top_color": NAVY, "bottom_color": NAVY,
        "left_color": NAVY, "right_color": NAVY,
    })
    tbl_row_fmt = workbook.add_format({
        "font_size": 10, "font_color": TEXT_PRIMARY,
        "font_name": "Calibri", "valign": "vcenter",
        "left": 1, "right": 1, "bottom": 1,
        "left_color": SUBTLE_BORDER, "right_color": SUBTLE_BORDER,
        "bottom_color": SUBTLE_BORDER,
    })
    tbl_row_alt_fmt = workbook.add_format({
        "font_size": 10, "font_color": TEXT_PRIMARY,
        "font_name": "Calibri", "valign": "vcenter",
        "bg_color": "#F1F5F9",
        "left": 1, "right": 1, "bottom": 1,
        "left_color": SUBTLE_BORDER, "right_color": SUBTLE_BORDER,
        "bottom_color": SUBTLE_BORDER,
    })
    tbl_num_fmt = workbook.add_format({
        "font_size": 10, "font_color": TEXT_PRIMARY,
        "font_name": "Calibri", "valign": "vcenter",
        "align": "center", "num_format": "#,##0",
        "left": 1, "right": 1, "bottom": 1,
        "left_color": SUBTLE_BORDER, "right_color": SUBTLE_BORDER,
        "bottom_color": SUBTLE_BORDER,
    })
    tbl_num_alt_fmt = workbook.add_format({
        "font_size": 10, "font_color": TEXT_PRIMARY,
        "font_name": "Calibri", "valign": "vcenter",
        "align": "center", "num_format": "#,##0",
        "bg_color": "#F1F5F9",
        "left": 1, "right": 1, "bottom": 1,
        "left_color": SUBTLE_BORDER, "right_color": SUBTLE_BORDER,
        "bottom_color": SUBTLE_BORDER,
    })

    # Tier badge formats
    tier_fmts = {
        1: workbook.add_format({
            "bold": True, "font_size": 10, "font_color": "#92400E",
            "bg_color": "#FEF3C7", "align": "center", "valign": "vcenter",
            "left": 1, "right": 1, "bottom": 1,
            "left_color": SUBTLE_BORDER, "right_color": SUBTLE_BORDER,
            "bottom_color": SUBTLE_BORDER,
        }),
        2: workbook.add_format({
            "bold": True, "font_size": 10, "font_color": "#1E40AF",
            "bg_color": "#DBEAFE", "align": "center", "valign": "vcenter",
            "left": 1, "right": 1, "bottom": 1,
            "left_color": SUBTLE_BORDER, "right_color": SUBTLE_BORDER,
            "bottom_color": SUBTLE_BORDER,
        }),
        3: workbook.add_format({
            "bold": True, "font_size": 10, "font_color": "#065F46",
            "bg_color": "#D1FAE5", "align": "center", "valign": "vcenter",
            "left": 1, "right": 1, "bottom": 1,
            "left_color": SUBTLE_BORDER, "right_color": SUBTLE_BORDER,
            "bottom_color": SUBTLE_BORDER,
        }),
        4: workbook.add_format({
            "bold": True, "font_size": 10, "font_color": "#374151",
            "bg_color": "#F3F4F6", "align": "center", "valign": "vcenter",
            "left": 1, "right": 1, "bottom": 1,
            "left_color": SUBTLE_BORDER, "right_color": SUBTLE_BORDER,
            "bottom_color": SUBTLE_BORDER,
        }),
    }

    # ── Column widths ─────────────────────────────────────────────────
    ws.set_column("A:A", 3, bg_fmt)   # left gutter
    ws.set_column("B:B", 22, bg_fmt)
    ws.set_column("C:C", 12, bg_fmt)
    ws.set_column("D:D", 3, bg_fmt)   # gap
    ws.set_column("E:E", 22, bg_fmt)
    ws.set_column("F:F", 12, bg_fmt)
    ws.set_column("G:G", 3, bg_fmt)   # gap
    ws.set_column("H:H", 22, bg_fmt)
    ws.set_column("I:I", 12, bg_fmt)
    ws.set_column("J:J", 3, bg_fmt)   # gap
    ws.set_column("K:K", 22, bg_fmt)
    ws.set_column("L:L", 12, bg_fmt)
    ws.set_column("M:M", 3, bg_fmt)   # right gutter

    ws.hide_gridlines(2)
    ws.set_tab_color(ACCENT)

    # ── Compute data ──────────────────────────────────────────────────
    region_counts = Counter(j.get("region", "Other") for j in all_jobs)
    tier_counts = Counter(j.get("tier") for j in all_jobs if j.get("tier"))
    field_counts = Counter(
        j.get("field", "Unknown") for j in all_jobs if j.get("field")
    )
    top_fields = field_counts.most_common(10)

    # Top institutions by job count
    inst_counts = Counter(
        j.get("institute", "Unknown") for j in all_jobs
        if j.get("institute") and j["institute"].lower() not in _AGGREGATOR_INSTITUTES
    )
    top_institutions = inst_counts.most_common(10)

    # Deadline urgency buckets
    urgent_count = 0    # ≤7 days
    soon_count = 0      # 8-30 days
    later_count = 0     # >30 days
    no_deadline = 0
    for j in all_jobs:
        dl = j.get("deadline") or ""
        try:
            days_left = (date.fromisoformat(dl) - date.today()).days
            if days_left < 0:
                continue
            elif days_left <= 7:
                urgent_count += 1
            elif days_left <= 30:
                soon_count += 1
            else:
                later_count += 1
        except (ValueError, TypeError):
            no_deadline += 1

    us_count = region_counts.get("US", 0)
    eu_count = region_counts.get("EU", 0)
    korea_count = region_counts.get("Korea", 0)
    asia_count = region_counts.get("Asia", 0)
    other_count = region_counts.get("Other", 0)

    # ══════════════════════════════════════════════════════════════════
    # ROW 0-1: HEADER
    # ══════════════════════════════════════════════════════════════════
    row = 1
    ws.merge_range(row, 1, row, 11,
                   "Job Search Pipeline", header_fmt)
    row += 1
    ws.merge_range(row, 1, row, 11,
                   f"Dashboard generated {datetime.now().strftime('%B %d, %Y at %H:%M')}   |   "
                   f"{len(all_jobs)} total positions tracked   |   "
                   f"{len(rec_pis)} PI recommendations",
                   subtitle_fmt)

    # ══════════════════════════════════════════════════════════════════
    # ROW 3-5: KPI CARDS
    # ══════════════════════════════════════════════════════════════════
    row = 4
    kpi_slate = _make_kpi_card(SLATE)
    kpi_data = [
        (1,  kpi_blue,    len(all_jobs), "Total Positions"),
        (4,  kpi_gold,    us_count,      "US Positions"),
        (7,  kpi_emerald, eu_count,      "EU Positions"),
        (10, kpi_slate,   korea_count,   "Korea Positions"),
    ]

    for col_start, (top_f, big_f, lbl_f), value, label in kpi_data:
        ws.merge_range(row, col_start, row, col_start + 1, "", top_f)
        ws.merge_range(row + 1, col_start, row + 1, col_start + 1, value, big_f)
        ws.merge_range(row + 2, col_start, row + 2, col_start + 1, label, lbl_f)

    # ══════════════════════════════════════════════════════════════════
    # ROW 8-9: SECONDARY KPI CARDS
    # ══════════════════════════════════════════════════════════════════
    row = 8

    kpi_navy = _make_kpi_card(NAVY)
    kpi_accent = _make_kpi_card(ACCENT_DARK)

    t1_count = tier_counts.get(1, 0) + tier_counts.get(2, 0)
    t3_count = tier_counts.get(3, 0) + tier_counts.get(4, 0)

    secondary_kpis = [
        (1,  kpi_navy,    len(rec_pis),   "PI Recommendations"),
        (4,  kpi_accent,  t1_count,       "Tier 1-2 Positions"),
        (7,  kpi_emerald, t3_count,       "Tier 3-4 Positions"),
        (10, kpi_rose,    urgent_count,   "Urgent (≤7 days)"),
    ]

    for col_start, (top_f, big_f, lbl_f), value, label in secondary_kpis:
        ws.merge_range(row, col_start, row, col_start + 1, "", top_f)
        ws.merge_range(row + 1, col_start, row + 1, col_start + 1, value, big_f)
        ws.merge_range(row + 2, col_start, row + 2, col_start + 1, label, lbl_f)

    # ══════════════════════════════════════════════════════════════════
    # ROW 12: REGION DATA (hidden, for chart) + REGION PIE CHART
    # ══════════════════════════════════════════════════════════════════
    chart_data_row = 12
    ws.merge_range(chart_data_row, 1, chart_data_row, 2,
                   "Region Breakdown", section_title_fmt)
    chart_data_row += 1

    region_data = [
        ("US", us_count),
        ("EU", eu_count),
        ("Korea", korea_count),
        ("Asia", asia_count),
        ("Other", other_count),
    ]
    region_colors = ["#2563EB", "#16A34A", "#9333EA", "#EA580C", "#9CA3AF"]

    ws.write(chart_data_row, 1, "Region", tbl_header_fmt)
    ws.write(chart_data_row, 2, "Count", tbl_header_fmt)
    chart_data_row += 1
    region_start = chart_data_row
    for i, (region, count) in enumerate(region_data):
        row_f = tbl_row_alt_fmt if i % 2 else tbl_row_fmt
        num_f = tbl_num_alt_fmt if i % 2 else tbl_num_fmt
        ws.write(chart_data_row, 1, region, row_f)
        ws.write(chart_data_row, 2, count, num_f)
        chart_data_row += 1
    region_end = chart_data_row - 1

    if any(c > 0 for _, c in region_data):
        chart = workbook.add_chart({"type": "doughnut"})
        chart.add_series({
            "name": "Jobs by Region",
            "categories": ["Summary Dashboard", region_start, 1, region_end, 1],
            "values": ["Summary Dashboard", region_start, 2, region_end, 2],
            "data_labels": {"percentage": True, "category": True,
                            "font": {"name": "Calibri", "size": 10, "color": TEXT_PRIMARY}},
            "points": [{"fill": {"color": c}} for c in region_colors],
        })
        chart.set_title({"name": "Jobs by Region",
                         "name_font": {"name": "Calibri", "size": 12,
                                       "color": NAVY, "bold": True}})
        chart.set_size({"width": 530, "height": 280})
        chart.set_legend({"position": "bottom",
                          "font": {"name": "Calibri", "size": 9}})
        chart.set_chartarea({"fill": {"color": LIGHT_BG}, "border": {"none": True}})
        chart.set_plotarea({"fill": {"color": LIGHT_BG}, "border": {"none": True}})
        # Position one row below section title (row 13, 0-based), col E (4)
        ws.insert_chart(13, 4, chart, {"x_offset": 0, "y_offset": 0})

    # ══════════════════════════════════════════════════════════════════
    # ROW 12: TIER BREAKDOWN TABLE (right side)
    # ══════════════════════════════════════════════════════════════════
    tier_row = 12
    ws.merge_range(tier_row, 10, tier_row, 11,
                   "Tier Distribution", section_title_fmt)
    tier_row += 1
    ws.write(tier_row, 10, "Tier", tbl_header_fmt)
    ws.write(tier_row, 11, "Count", tbl_header_fmt)
    tier_row += 1

    tier_labels = [
        (1, "Tier 1 (Top 50)"),
        (2, "Tier 2 (51-200)"),
        (3, "Tier 3 (201-500)"),
        (4, "Tier 4 (501+)"),
    ]
    tier_data_start = tier_row
    for i, (t, label) in enumerate(tier_labels):
        count = tier_counts.get(t, 0)
        t_fmt = tier_fmts.get(t, tbl_row_fmt)
        num_f = tbl_num_alt_fmt if i % 2 else tbl_num_fmt
        ws.write(tier_row, 10, label, t_fmt)
        ws.write(tier_row, 11, count, num_f)
        tier_row += 1

    # Unranked row
    ranked_total = sum(tier_counts.get(t, 0) for t in (1, 2, 3, 4))
    unranked = len(all_jobs) - ranked_total
    ws.write(tier_row, 10, "Unranked", tbl_row_alt_fmt)
    ws.write(tier_row, 11, unranked, tbl_num_alt_fmt)
    tier_row += 1

    # Deadline urgency mini-table
    tier_row += 1
    ws.merge_range(tier_row, 10, tier_row, 11,
                   "Deadline Urgency", section_title_fmt)
    tier_row += 1
    ws.write(tier_row, 10, "Window", tbl_header_fmt)
    ws.write(tier_row, 11, "Count", tbl_header_fmt)
    tier_row += 1

    urgency_data = [
        ("≤ 7 days (Urgent)", urgent_count),
        ("8-30 days (Soon)", soon_count),
        ("> 30 days", later_count),
        ("No deadline", no_deadline),
    ]

    urgent_fmt = workbook.add_format({
        "bold": True, "font_size": 10, "font_color": "#991B1B",
        "bg_color": "#FEE2E2", "valign": "vcenter",
        "left": 1, "right": 1, "bottom": 1,
        "left_color": SUBTLE_BORDER, "right_color": SUBTLE_BORDER,
        "bottom_color": SUBTLE_BORDER,
    })
    soon_fmt = workbook.add_format({
        "bold": True, "font_size": 10, "font_color": "#92400E",
        "bg_color": "#FEF3C7", "valign": "vcenter",
        "left": 1, "right": 1, "bottom": 1,
        "left_color": SUBTLE_BORDER, "right_color": SUBTLE_BORDER,
        "bottom_color": SUBTLE_BORDER,
    })
    urgency_fmts = [urgent_fmt, soon_fmt, tbl_row_fmt, tbl_row_alt_fmt]
    for i, ((label, count), uf) in enumerate(zip(urgency_data, urgency_fmts)):
        num_f = tbl_num_alt_fmt if i % 2 else tbl_num_fmt
        ws.write(tier_row, 10, label, uf)
        ws.write(tier_row, 11, count, num_f)
        tier_row += 1

    # ══════════════════════════════════════════════════════════════════
    # ROW ~20: TOP FIELDS BAR CHART + TABLE
    # ══════════════════════════════════════════════════════════════════
    fields_row = max(chart_data_row, tier_row) + 2
    ws.merge_range(fields_row, 1, fields_row, 2,
                   "Top Research Fields", section_title_fmt)
    fields_row += 1

    ws.write(fields_row, 1, "Field", tbl_header_fmt)
    ws.write(fields_row, 2, "Count", tbl_header_fmt)
    fields_row += 1
    field_start = fields_row
    for i, (field, count) in enumerate(top_fields):
        row_f = tbl_row_alt_fmt if i % 2 else tbl_row_fmt
        num_f = tbl_num_alt_fmt if i % 2 else tbl_num_fmt
        ws.write(fields_row, 1, field[:40], row_f)
        ws.write(fields_row, 2, count, num_f)
        fields_row += 1
    field_end = fields_row - 1

    # Save the section title row for chart positioning
    fields_section_row = fields_row - len(top_fields) - 2

    if top_fields:
        chart2 = workbook.add_chart({"type": "bar"})
        vivid_blues = [
            "#0037B3", "#0044CC", "#0055E6", "#0066FF", "#1A75FF",
            "#3385FF", "#4D94FF", "#66A3FF", "#80B3FF", "#99C2FF",
        ]
        chart2.add_series({
            "name": "Jobs by Field",
            "categories": ["Summary Dashboard", field_start, 1, field_end, 1],
            "values": ["Summary Dashboard", field_start, 2, field_end, 2],
            "points": [{"fill": {"color": vivid_blues[i % len(vivid_blues)]}}
                       for i in range(len(top_fields))],
            "gap": 80,
        })
        chart2.set_title({"name": "Top Research Fields",
                          "name_font": {"name": "Calibri", "size": 12,
                                        "color": NAVY, "bold": True}})
        chart2.set_size({"width": 530, "height": 320})
        chart2.set_legend({"none": True})
        chart2.set_chartarea({"fill": {"color": LIGHT_BG}, "border": {"none": True}})
        chart2.set_plotarea({"fill": {"color": LIGHT_BG}, "border": {"none": True}})
        chart2.set_y_axis({"num_font": {"name": "Calibri", "size": 9, "color": TEXT_SECONDARY}})
        chart2.set_x_axis({"num_font": {"name": "Calibri", "size": 9, "color": TEXT_SECONDARY},
                           "major_gridlines": {"visible": True, "line": {"color": SUBTLE_BORDER}}})
        # Position one row below section title, col E
        ws.insert_chart(fields_section_row + 1, 4, chart2, {"x_offset": 0, "y_offset": 0})

    # ══════════════════════════════════════════════════════════════════
    # TOP INSTITUTIONS TABLE (right of fields chart)
    # ══════════════════════════════════════════════════════════════════
    inst_row = fields_row - len(top_fields) - 2  # align with fields section
    ws.merge_range(inst_row, 10, inst_row, 11,
                   "Top Institutions", section_title_fmt)
    inst_row += 1

    ws.write(inst_row, 10, "Institution", tbl_header_fmt)
    ws.write(inst_row, 11, "Jobs", tbl_header_fmt)
    inst_row += 1
    for i, (inst, count) in enumerate(top_institutions):
        row_f = tbl_row_alt_fmt if i % 2 else tbl_row_fmt
        num_f = tbl_num_alt_fmt if i % 2 else tbl_num_fmt
        ws.write(inst_row, 10, inst[:35], row_f)
        ws.write(inst_row, 11, count, num_f)
        inst_row += 1

    # ══════════════════════════════════════════════════════════════════
    # FOOTER
    # ══════════════════════════════════════════════════════════════════
    footer_row = max(fields_row, inst_row) + 2
    footer_fmt = workbook.add_format({
        "font_size": 9, "font_color": TEXT_MUTED,
        "font_name": "Calibri", "bg_color": LIGHT_BG,
        "top": 1, "top_color": SUBTLE_BORDER,
    })
    ws.merge_range(footer_row, 1, footer_row, 11,
                   f"Generated by Job Search Pipeline   |   "
                   f"Data covers {len(all_jobs)} positions across "
                   f"{len(set(j.get('country') for j in all_jobs if j.get('country')))} countries   |   "
                   f"{datetime.now().strftime('%Y-%m-%d %H:%M')}",
                   footer_fmt)

    # Fill background for empty areas
    for r in range(footer_row + 1):
        for empty_col in (0, 3, 6, 9, 12):
            if empty_col < 13:
                ws.write_blank(r, empty_col, "", bg_fmt)


def _is_excluded(job: dict) -> bool:
    """Return True if the job matches any EXCLUDE_KEYWORDS or EXCLUDE_TITLE_KEYWORDS.

    Keyword exclusion uses title + field + keywords only (NOT description),
    because broad keywords like 'education' or 'history' cause massive
    false positives when matched against free-text descriptions.
    """
    # Expired deadline: remove jobs whose deadline has passed
    dl = job.get("deadline") or ""
    if dl:
        try:
            if date.fromisoformat(dl) < date.today():
                return True
        except (ValueError, TypeError):
            pass

    # Past-year titles: "2025년", "2024년" etc. are clearly old postings
    title_raw = job.get("title") or ""
    current_year = date.today().year
    for past_yr in range(current_year - 2, current_year):
        if str(past_yr) in title_raw:
            return True

    # Stale postings: posted in a previous year with no future deadline
    posted = job.get("posted_date") or ""
    if posted and not dl:
        try:
            if date.fromisoformat(posted).year < current_year:
                return True
        except (ValueError, TypeError):
            pass

    # Field-level exclusion: match against title + field + keywords only
    # (NOT description — too many false positives from phrases like
    #  "PhD in biology required", "publication history", "by law")
    field_blob = " ".join(
        (job.get(k) or "") for k in ("title", "field", "keywords")
    ).lower()
    if any(kw.lower() in field_blob for kw in EXCLUDE_KEYWORDS):
        return True

    # PhD/doctoral exclusion: check title only (not description)
    title = (job.get("title") or "").lower()
    if any(kw.lower() in title for kw in _PHD_TITLE_KEYWORDS):
        # Exception: keep if title also says "postdoc" or "research fellow"
        if any(k in title for k in ("postdoc", "post-doc", "postdoctoral",
                                     "post-doctoral", "research fellow")):
            return False
        return True

    # Title-only exclusion (non-researcher positions)
    # Normalize Unicode quotes so "bachelor\u2019s" matches "bachelor's"
    title_norm = title.replace("\u2018", "'").replace("\u2019", "'")
    if any(kw.lower() in title_norm for kw in EXCLUDE_TITLE_KEYWORDS):
        # Exception: keep if title also says "postdoc" or "research fellow"
        if any(k in title for k in ("postdoc", "post-doc", "postdoctoral",
                                     "post-doctoral", "research fellow")):
            return False
        return True

    # Faculty exclusion: remove from US/EU/Other, keep for Korea
    if any(kw.lower() in title for kw in FACULTY_TITLE_KEYWORDS):
        if job.get("region") != "Korea":
            if not any(k in title for k in ("postdoc", "post-doc", "postdoctoral",
                                             "post-doctoral", "research fellow")):
                return True

    # Garbage detection: titles that are too short or clearly not job postings
    # For Korean titles, allow shorter (Korean chars are dense)
    min_len = 4 if _has_korean(title) else 10
    if len(title.strip()) < min_len:
        return True
    if any(pat.lower() in title for pat in GARBAGE_TITLE_PATTERNS):
        return True

    # Korean jobs: require bio/research signal in title.
    # Korean job boards (esp. HiBrainNet) list ALL types of positions
    # (banks, government, factories, etc.).  Without this check,
    # irrelevant jobs like 소상공인, 기념관장, 임기제공무원 leak through.
    if _has_korean(job.get("title") or ""):
        if not _is_bio_relevant_korean_title(job.get("title") or ""):
            return True

    return False


def _has_korean(text: str) -> bool:
    """Return True if text contains Korean characters."""
    return bool(re.search(r"[\uac00-\ud7a3]", text))


_KOREAN_BIO_TITLE_SIGNALS = [
    # Postdoc positions (always research-relevant)
    "박사후", "박사 후", "석사후", "석사 후",
    "postdoc", "post-doc", "postdoctoral",
    # Bio / medical / science fields (Korean)
    "바이오", "생명", "생물", "약학", "의과", "의학", "의생명",
    "뇌과학", "뇌공학", "유전", "분자", "단백질", "효소", "미생물",
    "세포", "게놈", "면역", "약리", "생화학", "발효", "보건연구",
    "생명공학", "합성생물", "대사공학", "식품공학",
    # English bio terms
    "biotech", "crispr", "genomic", "protein", "biology",
    "molecular", "biochem", "bioinformat", "microbio",
    "synthetic bio", "enzyme", "ferment",
]


def _is_bio_relevant_korean_title(title: str) -> bool:
    """Return True if a Korean job title contains bio/research signals."""
    t = title.lower()
    return any(sig in t for sig in _KOREAN_BIO_TITLE_SIGNALS)


# PhD/Doctoral keywords — matched against TITLE only (not description)
_PHD_TITLE_KEYWORDS = [
    "PhD position",
    "PhD student",
    "PhD fellow",
    "doctoral position",
    "doctoral student",
    "doctoral researcher",
    "doctoral fellow",
    "doctoral grant",
    "PhD candidate",
    "doctoral candidate",
    "PhD fellowship",
]


def _is_faculty_position(job: dict) -> bool:
    """Return True if the job title matches a faculty/professor keyword."""
    title = (job.get("title") or "").lower()
    return any(kw.lower() in title for kw in FACULTY_TITLE_KEYWORDS)


def _build_institute_rank() -> dict[str, int]:
    """Build lowercase institute -> ranking position within its tier.

    Lower number = higher ranked.  Institutions not in the rankings
    get a large default value.
    """
    from src.config import load_rankings
    rankings = load_rankings()
    rank: dict[str, int] = {}
    pos = 0
    for tier_key in ("1", "2", "3", "4"):
        info = rankings.get("tiers", {}).get(tier_key, {})
        for inst in info.get("institutions", []):
            rank[inst.lower()] = pos
            pos += 1
    # Companies
    companies = rankings.get("companies", {})
    for section in ("top_companies", "companies"):
        sec = companies.get(section, {})
        insts = sec.get("institutions", []) if isinstance(sec, dict) else sec
        for inst in insts:
            rank[inst.lower()] = pos
            pos += 1
    # Aliases
    aliases = rankings.get("tier_lookup_aliases", {})
    for alias, canonical in aliases.items():
        if canonical.lower() in rank and alias.lower() not in rank:
            rank[alias.lower()] = rank[canonical.lower()]
    return rank


_INSTITUTE_RANK: dict[str, int] | None = None


def _get_institute_rank(institute: str) -> int:
    """Return the ranking position for an institute (lower = better)."""
    global _INSTITUTE_RANK
    if _INSTITUTE_RANK is None:
        _INSTITUTE_RANK = _build_institute_rank()
    key = institute.strip().lower()
    if key in _INSTITUTE_RANK:
        return _INSTITUTE_RANK[key]
    # Partial match
    for name, pos in _INSTITUTE_RANK.items():
        if name in key or key in name:
            return pos
    return 99999


def _sort_by_tier(jobs: list[dict]) -> list[dict]:
    """Sort jobs by tier, then by institution ranking within each tier.

    Sort order: Company → T1 → T2 → T3 → T4 → Unranked
    Within each tier: institution ranking order (descending).
    """
    def _key(j: dict):
        tier = j.get("tier")
        tier_num = tier if isinstance(tier, int) else 999
        company = is_company(j.get("institute") or "")
        # Companies sort at the top (before T1)
        if company:
            sort_tier = 0
        else:
            sort_tier = tier_num
        inst_rank = _get_institute_rank(j.get("institute") or "")
        return (sort_tier, inst_rank)
    return sorted(jobs, key=_key)


def _sort_pis_by_tier(pis: list[dict]) -> list[dict]:
    """Sort recommended PIs: Company first, then by tier and institution ranking.

    Sort order: Company → T1 → T2 → T3 → T4 → Unranked (no tier)
    Within each tier: institution ranking order.
    """
    def _key(p: dict):
        tier = p.get("tier")
        tier_num = tier if isinstance(tier, int) and 1 <= tier <= 4 else 999
        company = is_company(p.get("institute") or "")
        if company:
            sort_tier = 0
        else:
            sort_tier = tier_num
        inst_rank = _get_institute_rank(p.get("institute") or "")
        return (sort_tier, inst_rank)
    return sorted(pis, key=_key)


def _find_previous_excel(output_dir: Path) -> Path | None:
    """Find the most recent previous Excel file to preserve user edits."""
    import glob as _glob

    # Fixed name first
    fixed = output_dir / "JobSearch_Auto.xlsx"
    if fixed.exists():
        return fixed
    # Fallback: dated files
    pattern = str(output_dir / "JobSearch_Auto_*.xlsx")
    files = sorted(_glob.glob(pattern), key=lambda f: Path(f).stat().st_mtime, reverse=True)
    return Path(files[0]) if files else None


def _read_previous_excel(output_dir: Path) -> dict[str, dict[str, pd.DataFrame]]:
    """Read previous Excel into DataFrames keyed by sheet name.

    Returns {sheet_name: DataFrame} for job sheets, preserving all user edits.
    Returns empty dict if no previous file exists.
    """
    prev_path = _find_previous_excel(output_dir)
    if not prev_path:
        return {}

    sheets: dict[str, pd.DataFrame] = {}
    for sheet in ("US Positions", "EU Positions", "Korea Positions", "Other Positions"):
        try:
            df = pd.read_excel(str(prev_path), sheet_name=sheet, engine="openpyxl")
            sheets[sheet] = df
        except Exception:
            continue
    return sheets


def _detect_dismissed_urls(
    prev_sheets: dict[str, pd.DataFrame],
    exportable_urls: set[str],
) -> set[str]:
    """Detect URLs that the user removed from the previous Excel.

    Compares previously exported URLs (from DB exported_at) against
    what's still in the Excel. Missing URLs = user-dismissed.
    Records them in the dismissed_urls table.
    """
    from src.db import get_connection, dismiss_urls

    # Collect all Job URLs currently in the Excel
    prev_urls: set[str] = set()
    for df in prev_sheets.values():
        if "Job URL" not in df.columns:
            continue
        for url in df["Job URL"].dropna():
            url = str(url).strip()
            if url and url != "nan" and url.startswith("http"):
                prev_urls.add(url)

    if not prev_urls:
        return set()

    # Sync user status changes from Excel → DB
    with get_connection() as conn:
        for df in prev_sheets.values():
            if "Job URL" not in df.columns or "Status" not in df.columns:
                continue
            for _, row in df.iterrows():
                url = str(row.get("Job URL", ""))
                status = row.get("Status", "")
                if url and url.startswith("http") and pd.notna(status):
                    status = str(status).strip()
                    if status and status not in ("new", "nan"):
                        conn.execute(
                            "UPDATE jobs SET status = ? WHERE url = ?",
                            (status, url),
                        )

    # Detect dismissed: previously exported but removed from Excel
    dismissed: set[str] = set()
    with get_connection() as conn:
        exported_rows = conn.execute(
            "SELECT url FROM jobs WHERE exported_at IS NOT NULL "
            "AND (status IS NULL OR status NOT IN ('dismissed'))"
        ).fetchall()

        for row in exported_rows:
            url = row["url"]
            if url in exportable_urls and url not in prev_urls:
                dismissed.add(url)

    # Persist to dismissed_urls table (permanent blacklist)
    if dismissed:
        dismiss_urls(list(dismissed))
        logger.info("Detected %d user-dismissed jobs from previous Excel", len(dismissed))

    return dismissed


def _merge_with_previous(
    prev_sheets: dict[str, pd.DataFrame],
    new_jobs_by_region: dict[str, list[dict]],
    dismissed_urls: set[str],
) -> dict[str, tuple[pd.DataFrame, list[dict]]]:
    """Merge previous Excel data with new jobs.

    Preserves all existing rows from the previous Excel (user edits intact).
    Only adds genuinely new jobs (URL not already present).
    Returns {sheet_name: (merged_df, new_job_dicts)} for sorting.
    """
    sheet_region_map = {
        "US Positions": "US",
        "EU Positions": "EU",
        "Korea Positions": "Korea",
        "Other Positions": "Other",
    }

    result: dict[str, tuple[pd.DataFrame, list[dict]]] = {}

    for sheet_name, region_key in sheet_region_map.items():
        prev_df = prev_sheets.get(sheet_name)
        new_jobs = new_jobs_by_region.get(region_key, [])

        # Collect existing URLs from previous Excel
        existing_urls: set[str] = set()
        if prev_df is not None and "Job URL" in prev_df.columns:
            for url in prev_df["Job URL"].dropna():
                url_str = str(url).strip()
                if url_str and url_str != "nan":
                    existing_urls.add(url_str)

        # Filter new jobs: only those not already in Excel and not dismissed
        genuinely_new = [
            j for j in new_jobs
            if j.get("url")
            and j["url"] not in existing_urls
            and j["url"] not in dismissed_urls
        ]

        if prev_df is not None and len(prev_df) > 0:
            # Remove dismissed URLs from previous data too
            if dismissed_urls and "Job URL" in prev_df.columns:
                prev_df = prev_df[~prev_df["Job URL"].isin(dismissed_urls)]

            # Ensure columns match
            if genuinely_new:
                new_rows_df = pd.DataFrame(
                    [_job_to_row(j) for j in genuinely_new], columns=JOB_COLUMNS
                )
                # Align columns: use the union of both, preserving prev columns
                all_cols = list(prev_df.columns)
                for col in new_rows_df.columns:
                    if col not in all_cols:
                        all_cols.append(col)

                # Reindex both to same columns
                prev_df = prev_df.reindex(columns=all_cols, fill_value="")
                new_rows_df = new_rows_df.reindex(columns=all_cols, fill_value="")
                merged_df = pd.concat([prev_df, new_rows_df], ignore_index=True)
            else:
                merged_df = prev_df
        elif genuinely_new:
            merged_df = pd.DataFrame(
                [_job_to_row(j) for j in genuinely_new], columns=JOB_COLUMNS
            )
        else:
            merged_df = pd.DataFrame(columns=JOB_COLUMNS)

        result[sheet_name] = (merged_df, genuinely_new)

    return result


def _sort_merged_df(df: pd.DataFrame) -> pd.DataFrame:
    """Sort a merged DataFrame by Tier then Institute ranking."""
    if df.empty:
        return df

    # Sort by: Company first → T1 → T2 → T3 → T4 → Unranked
    # Within each tier: by institute ranking
    def _tier_sort_key(tier_val):
        tier_str = str(tier_val).strip() if pd.notna(tier_val) else ""
        if tier_str == "Company":
            return 0
        if tier_str.startswith("T") and len(tier_str) == 2 and tier_str[1].isdigit():
            return int(tier_str[1])
        return 99

    def _inst_sort_key(inst_val):
        inst = str(inst_val).strip() if pd.notna(inst_val) else ""
        return _get_institute_rank(inst)

    if "Tier" in df.columns and "Institute" in df.columns:
        df = df.assign(
            _tier_sort=df["Tier"].apply(_tier_sort_key),
            _inst_sort=df["Institute"].apply(_inst_sort_key),
        )
        df = df.sort_values(["_tier_sort", "_inst_sort"], ascending=True)
        df = df.drop(columns=["_tier_sort", "_inst_sort"])
        df = df.reset_index(drop=True)

    return df


def _mark_exported(urls: list[str]) -> None:
    """Set exported_at timestamp for all exported job URLs."""
    if not urls:
        return
    from src.db import get_connection
    with get_connection() as conn:
        conn.executemany(
            "UPDATE jobs SET exported_at = datetime('now') WHERE url = ?",
            [(u,) for u in urls],
        )


# ---------------------------------------------------------------------------
# openpyxl incremental update helpers
# ---------------------------------------------------------------------------

_HEADER_WIDTH_MAP = {
    "Title": 65, "PI Name": 18, "Institute": 25, "Department": 20,
    "Tier": 5, "Country": 12, "Region": 6,
    "Field": 20, "Keywords": 35,
    "Position Summary": 45, "Responsibilities": 45, "Preferred Qualifications": 40,
    "Degree Required": 30, "Skills/Techniques": 30, "Requirements (Full)": 40,
    "Salary": 18, "Duration": 18, "Contract Type": 12, "Start Date": 14,
    "Conditions (Full)": 30,
    "Posted Date": 11, "Deadline": 11,
    "Application Materials": 30,
    "Description": 50,
    **{f"Recent Paper {i+1}": 35 for i in range(_MAX_PAPER_COLS)},
    **{f"Top Cited Paper {i+1}": 35 for i in range(_MAX_PAPER_COLS)},
    "Contact Email": 25,
    **{f"Info URL {i+1}": 20 for i in range(_MAX_INFO_URLS)},
    "Job URL": 15, "Job URL 2": 15, "Lab URL": 15, "Scholar URL": 15, "Dept URL": 15,
    "Match Score": 8, "Source": 12, "Status": 7,
}

_TIER_COLORS = {
    "T1": {"bg": "FFF3CD", "fg": "856404", "bold": True},
    "T2": {"bg": "CCE5FF", "fg": "004085", "bold": False},
    "T3": {"bg": "D4EDDA", "fg": "155724", "bold": False},
    "T4": {"bg": "E2E3E5", "fg": "383D41", "bold": False},
    "Company": {"bg": "E8DAEF", "fg": "6C3483", "bold": True},
}


def _find_column_index(ws, col_name: str) -> int | None:
    """Find 1-based column index by header name in row 1."""
    for col in range(1, (ws.max_column or 0) + 1):
        if ws.cell(row=1, column=col).value == col_name:
            return col
    return None


def _write_openpyxl_headers(ws) -> None:
    """Write JOB_COLUMNS as header row with formatting to a new openpyxl sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="16213e")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(JOB_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    for col_idx, col_name in enumerate(JOB_COLUMNS, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _HEADER_WIDTH_MAP.get(col_name, 15)

    ws.freeze_panes = "A2"


def _apply_tier_format_openpyxl(cell, tier_val: str) -> None:
    """Apply tier-based color to an openpyxl cell."""
    from openpyxl.styles import Font, PatternFill

    fmt = _TIER_COLORS.get(str(tier_val).strip())
    if fmt:
        cell.fill = PatternFill("solid", fgColor=fmt["bg"])
        cell.font = Font(color=fmt["fg"], bold=fmt["bold"])


def _write_paper_links_openpyxl(ws, row_idx: int, job: dict) -> None:
    """Write paper hyperlinks for a single row in an openpyxl sheet."""
    from openpyxl.styles import Font

    link_font = Font(color="0563C1", underline="single", size=9)

    for prefix, key in [("Recent Paper", "recent_papers"),
                        ("Top Cited Paper", "top_cited_papers")]:
        papers = _parse_papers(job.get(key))
        for i in range(min(_MAX_PAPER_COLS, len(papers))):
            col_name = f"{prefix} {i + 1}"
            col_idx = _find_column_index(ws, col_name)
            if not col_idx:
                continue
            p = papers[i]
            url = p.get("url", "")
            year = p.get("year") or "?"
            title = (p.get("title") or "Untitled")[:80]
            cites = p.get("citation_count", 0)
            display = f"({year}) {title} [{cites} cites]"
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = display
            if url:
                cell.hyperlink = url
                cell.font = link_font


def _style_new_row_openpyxl(ws, row_idx: int, job: dict) -> None:
    """Apply formatting to a newly appended row in an openpyxl sheet."""
    from openpyxl.styles import Font, PatternFill

    # Tier formatting
    tier_col = _find_column_index(ws, "Tier")
    if tier_col:
        tier_val = ws.cell(row=row_idx, column=tier_col).value
        if tier_val:
            _apply_tier_format_openpyxl(
                ws.cell(row=row_idx, column=tier_col), tier_val)

    # Deadline gradient
    dl_col = _find_column_index(ws, "Deadline")
    if dl_col:
        dl_val = ws.cell(row=row_idx, column=dl_col).value
        if dl_val:
            color, days = _deadline_bg_color(str(dl_val))
            if color:
                ws.cell(row=row_idx, column=dl_col).fill = PatternFill(
                    "solid", fgColor=color[1:])
                if days < 0:
                    ws.cell(row=row_idx, column=dl_col).font = Font(
                        color="666666", size=9)
                elif days <= 7:
                    ws.cell(row=row_idx, column=dl_col).font = Font(
                        bold=True, color="FFFFFF", size=9)
                elif days <= 30:
                    ws.cell(row=row_idx, column=dl_col).font = Font(
                        color="FFFFFF", size=9)

    # URL hyperlinks
    link_font = Font(color="0563C1", underline="single")
    for url_col_name in ("Job URL", "Job URL 2", "Lab URL", "Scholar URL",
                         "Dept URL", "Info URL 1", "Info URL 2", "Info URL 3"):
        col_idx = _find_column_index(ws, url_col_name)
        if col_idx:
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).startswith("http"):
                ws.cell(row=row_idx, column=col_idx).hyperlink = str(val)
                ws.cell(row=row_idx, column=col_idx).font = link_font

    # Paper hyperlinks
    _write_paper_links_openpyxl(ws, row_idx, job)

    # Citizenship restriction → red Institute
    if _has_citizenship_restriction(job):
        inst_col = _find_column_index(ws, "Institute")
        if inst_col:
            ws.cell(row=row_idx, column=inst_col).font = Font(
                color="CC0000", bold=True)


def _update_sheet(
    wb, sheet_name: str, region_key: str,
    new_jobs_by_region: dict, dismissed_urls: set,
) -> list[dict]:
    """Incrementally update a data sheet: delete dismissed rows, append new jobs.

    Returns the list of genuinely new jobs that were appended.
    """
    from openpyxl.utils import get_column_letter

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        _write_openpyxl_headers(ws)
    else:
        ws = wb[sheet_name]

    # 1. Collect existing URLs and identify dismissed rows
    url_col_idx = _find_column_index(ws, "Job URL")
    rows_to_delete = []
    existing_urls: set[str] = set()

    if url_col_idx:
        for row_idx in range(2, (ws.max_row or 1) + 1):
            url_val = ws.cell(row=row_idx, column=url_col_idx).value
            if url_val:
                url_str = str(url_val).strip()
                if url_str in dismissed_urls:
                    rows_to_delete.append(row_idx)
                elif url_str.startswith("http"):
                    existing_urls.add(url_str)

    # Delete bottom→top to avoid index shifting
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)

    # 2. Filter genuinely new jobs
    region_jobs = new_jobs_by_region.get(region_key, [])
    new_jobs = [j for j in region_jobs
                if j.get("url") and j["url"] not in existing_urls
                and j["url"] not in dismissed_urls]
    new_jobs = _sort_by_tier(new_jobs)

    # 3. Append new rows with formatting
    for job in new_jobs:
        row_data = _job_to_row(job)
        row_idx = (ws.max_row or 1) + 1
        for col_idx, col_name in enumerate(JOB_COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx,
                    value=row_data.get(col_name, ""))
        _style_new_row_openpyxl(ws, row_idx, job)

    # 4. Update auto-filter range
    if ws.max_row and ws.max_row > 1:
        last_col = get_column_letter(len(JOB_COLUMNS))
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    return new_jobs


# ---------------------------------------------------------------------------
# openpyxl Summary Dashboard & PI sheet rebuilders
# ---------------------------------------------------------------------------

def _rebuild_summary_dashboard_openpyxl(
    wb, all_jobs: list[dict], rec_pis: list[dict],
) -> None:
    """Delete and recreate the Summary Dashboard using openpyxl."""
    from collections import Counter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import DoughnutChart, BarChart, Reference
    from openpyxl.utils import get_column_letter

    if "Summary Dashboard" in wb.sheetnames:
        wb.remove(wb["Summary Dashboard"])

    ws = wb.create_sheet("Summary Dashboard", 0)
    ws.sheet_view.showGridLines = False

    # ── Color palette ─────────────────────────────────────────────────
    NAVY = "0F1B2D"
    ACCENT = "3B82F6"
    ACCENT_DARK = "1E40AF"
    GOLD = "F59E0B"
    EMERALD = "10B981"
    ROSE = "F43F5E"
    SLATE = "64748B"
    LIGHT_BG = "F8FAFC"
    CARD_BG = "FFFFFF"
    SUBTLE_BORDER = "E2E8F0"
    TEXT_PRIMARY = "1E293B"
    TEXT_SECONDARY = "475569"
    TEXT_MUTED = "94A3B8"

    bg_fill = PatternFill("solid", fgColor=LIGHT_BG)
    card_fill = PatternFill("solid", fgColor=CARD_BG)
    border_side = Side(style="thin", color=SUBTLE_BORDER)

    # ── Column widths ─────────────────────────────────────────────────
    col_widths = [3, 22, 12, 3, 22, 12, 3, 22, 12, 3, 22, 12, 3]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_properties.tabColor = ACCENT

    # ── Helpers ───────────────────────────────────────────────────────
    def _write_merged(r1, c1, r2, c2, value, font=None, fill=None,
                      alignment=None, border=None):
        if r1 != r2 or c1 != c2:
            ws.merge_cells(start_row=r1, start_column=c1,
                           end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border

    def _kpi_card(row, col_start, value, label, accent_color):
        accent_side = Side(style="medium", color=accent_color)
        # Top accent bar
        _write_merged(row, col_start, row, col_start + 1, "",
                      font=Font(size=1), fill=card_fill,
                      border=Border(top=accent_side, left=border_side,
                                    right=border_side))
        # Big number
        _write_merged(row + 1, col_start, row + 1, col_start + 1, value,
                      font=Font(bold=True, size=28, color=accent_color,
                                name="Calibri"),
                      fill=card_fill,
                      alignment=Alignment(horizontal="center",
                                          vertical="center"),
                      border=Border(left=border_side, right=border_side))
        # Label
        _write_merged(row + 2, col_start, row + 2, col_start + 1, label,
                      font=Font(size=9, color=TEXT_SECONDARY,
                                name="Calibri"),
                      fill=card_fill,
                      alignment=Alignment(horizontal="center",
                                          vertical="top"),
                      border=Border(left=border_side, right=border_side,
                                    bottom=border_side))

    def _write_tbl_header(row, col, value):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _write_tbl_row(row, col_label, col_val, label, value, alt=False):
        alt_fill = PatternFill("solid", fgColor="F1F5F9") if alt else None
        cl = ws.cell(row=row, column=col_label, value=label)
        cl.font = Font(size=10, color=TEXT_PRIMARY, name="Calibri")
        cv = ws.cell(row=row, column=col_val, value=value)
        cv.font = Font(size=10, color=TEXT_PRIMARY, name="Calibri")
        cv.alignment = Alignment(horizontal="center")
        if alt_fill:
            cl.fill = alt_fill
            cv.fill = alt_fill

    # ── Compute data ──────────────────────────────────────────────────
    region_counts = Counter(j.get("region", "Other") for j in all_jobs)
    tier_counts = Counter(j.get("tier") for j in all_jobs if j.get("tier"))
    field_counts = Counter(
        j.get("field", "Unknown") for j in all_jobs if j.get("field"))
    top_fields = field_counts.most_common(10)
    inst_counts = Counter(
        j.get("institute", "Unknown") for j in all_jobs
        if j.get("institute")
        and j["institute"].lower() not in _AGGREGATOR_INSTITUTES)
    top_institutions = inst_counts.most_common(10)

    urgent_count = soon_count = later_count = no_deadline = 0
    for j in all_jobs:
        dl = j.get("deadline") or ""
        try:
            days_left = (date.fromisoformat(dl) - date.today()).days
            if days_left < 0:
                continue
            elif days_left <= 7:
                urgent_count += 1
            elif days_left <= 30:
                soon_count += 1
            else:
                later_count += 1
        except (ValueError, TypeError):
            no_deadline += 1

    us_count = region_counts.get("US", 0)
    eu_count = region_counts.get("EU", 0)
    korea_count = region_counts.get("Korea", 0)
    asia_count = region_counts.get("Asia", 0)
    other_count = region_counts.get("Other", 0)

    # ── Row heights ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 6       # top gutter
    ws.row_dimensions[2].height = 36      # title
    ws.row_dimensions[3].height = 18      # subtitle
    ws.row_dimensions[4].height = 10      # gap
    ws.row_dimensions[5].height = 6       # KPI accent bar
    ws.row_dimensions[6].height = 42      # KPI big number
    ws.row_dimensions[7].height = 18      # KPI label
    ws.row_dimensions[8].height = 10      # gap
    ws.row_dimensions[9].height = 6       # KPI accent bar
    ws.row_dimensions[10].height = 42     # KPI big number
    ws.row_dimensions[11].height = 18     # KPI label
    ws.row_dimensions[12].height = 10     # gap

    # ═══════════════════════════════════════════════════════════════════
    # HEADER (rows 2-3)
    # ═══════════════════════════════════════════════════════════════════
    _write_merged(2, 2, 2, 12, "Job Search Pipeline",
                  font=Font(bold=True, size=22, color=NAVY, name="Calibri"),
                  fill=bg_fill)
    _write_merged(
        3, 2, 3, 12,
        f"Dashboard generated {datetime.now().strftime('%B %d, %Y at %H:%M')}   |   "
        f"{len(all_jobs)} total positions tracked   |   "
        f"{len(rec_pis)} PI recommendations",
        font=Font(size=10, color=TEXT_MUTED, name="Calibri"),
        fill=bg_fill)

    # ═══════════════════════════════════════════════════════════════════
    # PRIMARY KPI CARDS (rows 5-7)
    # ═══════════════════════════════════════════════════════════════════
    _kpi_card(5, 2, len(all_jobs), "Total Positions", ACCENT)
    _kpi_card(5, 5, us_count, "US Positions", GOLD)
    _kpi_card(5, 8, eu_count, "EU Positions", EMERALD)
    _kpi_card(5, 11, korea_count, "Korea Positions", SLATE)

    # ═══════════════════════════════════════════════════════════════════
    # SECONDARY KPI CARDS (rows 9-11)
    # ═══════════════════════════════════════════════════════════════════
    t1_count = tier_counts.get(1, 0) + tier_counts.get(2, 0)
    t3_count = tier_counts.get(3, 0) + tier_counts.get(4, 0)
    _kpi_card(9, 2, len(rec_pis), "PI Recommendations", NAVY)
    _kpi_card(9, 5, t1_count, "Tier 1-2 Positions", ACCENT_DARK)
    _kpi_card(9, 8, t3_count, "Tier 3-4 Positions", EMERALD)
    _kpi_card(9, 11, urgent_count, "Urgent (≤7 days)", ROSE)

    # ═══════════════════════════════════════════════════════════════════
    # REGION TABLE + DOUGHNUT CHART (row 13+)
    # ═══════════════════════════════════════════════════════════════════
    section_font = Font(bold=True, size=13, color=NAVY, name="Calibri")
    section_border = Border(bottom=Side(style="medium", color=ACCENT))

    ws.row_dimensions[13].height = 22     # section title
    ws.row_dimensions[14].height = 20     # table header
    for r in range(15, 20):
        ws.row_dimensions[r].height = 18  # table rows

    _write_merged(13, 2, 13, 3, "Region Breakdown",
                  font=section_font, fill=bg_fill, border=section_border)
    _write_tbl_header(14, 2, "Region")
    _write_tbl_header(14, 3, "Count")

    region_data = [
        ("US", us_count), ("EU", eu_count), ("Korea", korea_count),
        ("Asia", asia_count), ("Other", other_count),
    ]
    for i, (region, count) in enumerate(region_data):
        _write_tbl_row(15 + i, 2, 3, region, count, alt=bool(i % 2))

    # Doughnut chart — spans cols E-I (2nd+3rd box)
    if any(c > 0 for _, c in region_data):
        from openpyxl.chart.series import DataPoint

        chart = DoughnutChart()
        chart.title = "Jobs by Region"
        chart.style = 10
        data_ref = Reference(ws, min_col=3, min_row=14, max_row=19)
        cats_ref = Reference(ws, min_col=2, min_row=15, max_row=19)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        region_colors = ["2563EB", "16A34A", "9333EA", "EA580C", "9CA3AF"]
        series = chart.series[0]
        for i, color in enumerate(region_colors):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color
            series.data_points.append(pt)

        chart.width = 14
        chart.height = 7
        ws.add_chart(chart, "E14")

    # ═══════════════════════════════════════════════════════════════════
    # TIER DISTRIBUTION TABLE (right side, cols K-L, row 13+)
    # ═══════════════════════════════════════════════════════════════════
    _write_merged(13, 11, 13, 12, "Tier Distribution",
                  font=section_font, fill=bg_fill, border=section_border)
    _write_tbl_header(14, 11, "Tier")
    _write_tbl_header(14, 12, "Count")

    tier_labels = [
        (1, "Tier 1 (Top 50)", "FEF3C7", "92400E"),
        (2, "Tier 2 (51-200)", "DBEAFE", "1E40AF"),
        (3, "Tier 3 (201-500)", "D1FAE5", "065F46"),
        (4, "Tier 4 (501+)", "F3F4F6", "374151"),
    ]
    for i, (t, label, bg, fg) in enumerate(tier_labels):
        row = 15 + i
        count = tier_counts.get(t, 0)
        cl = ws.cell(row=row, column=11, value=label)
        cl.font = Font(bold=True, size=10, color=fg, name="Calibri")
        cl.fill = PatternFill("solid", fgColor=bg)
        cl.alignment = Alignment(horizontal="center", vertical="center")
        cv = ws.cell(row=row, column=12, value=count)
        cv.font = Font(size=10, color=TEXT_PRIMARY, name="Calibri")
        cv.alignment = Alignment(horizontal="center")

    ranked_total = sum(tier_counts.get(t, 0) for t in (1, 2, 3, 4))
    unranked = len(all_jobs) - ranked_total
    cl = ws.cell(row=19, column=11, value="Unranked")
    cl.font = Font(size=10, color=TEXT_PRIMARY)
    cl.fill = PatternFill("solid", fgColor="F1F5F9")
    cv = ws.cell(row=19, column=12, value=unranked)
    cv.font = Font(size=10, color=TEXT_PRIMARY)
    cv.alignment = Alignment(horizontal="center")
    cv.fill = PatternFill("solid", fgColor="F1F5F9")

    # ── Deadline urgency mini-table (cols K-L) ────────────────────────
    ws.row_dimensions[20].height = 8      # gap
    ws.row_dimensions[21].height = 22     # section title
    ws.row_dimensions[22].height = 20     # table header
    for r in range(23, 27):
        ws.row_dimensions[r].height = 18  # table rows
    ws.row_dimensions[27].height = 10     # gap

    _write_merged(21, 11, 21, 12, "Deadline Urgency",
                  font=section_font, fill=bg_fill, border=section_border)
    _write_tbl_header(22, 11, "Window")
    _write_tbl_header(22, 12, "Count")

    urgency_data = [
        ("≤ 7 days (Urgent)", urgent_count, "FEE2E2", "991B1B", True),
        ("8-30 days (Soon)", soon_count, "FEF3C7", "92400E", True),
        ("> 30 days", later_count, None, TEXT_PRIMARY, False),
        ("No deadline", no_deadline, "F1F5F9", TEXT_PRIMARY, False),
    ]
    for i, (label, count, bg, fg, bold) in enumerate(urgency_data):
        row = 23 + i
        cl = ws.cell(row=row, column=11, value=label)
        cl.font = Font(bold=bold, size=10, color=fg)
        if bg:
            cl.fill = PatternFill("solid", fgColor=bg)
        cv = ws.cell(row=row, column=12, value=count)
        cv.font = Font(size=10, color=TEXT_PRIMARY)
        cv.alignment = Alignment(horizontal="center")

    # ═══════════════════════════════════════════════════════════════════
    # TOP FIELDS TABLE + BAR CHART (row 28+)
    # ═══════════════════════════════════════════════════════════════════
    fields_start = 28
    ws.row_dimensions[fields_start].height = 22      # section title
    ws.row_dimensions[fields_start + 1].height = 20  # table header
    for r in range(fields_start + 2, fields_start + 2 + max(len(top_fields), 1)):
        ws.row_dimensions[r].height = 18

    _write_merged(fields_start, 2, fields_start, 3, "Top Research Fields",
                  font=section_font, fill=bg_fill, border=section_border)
    _write_tbl_header(fields_start + 1, 2, "Field")
    _write_tbl_header(fields_start + 1, 3, "Count")

    for i, (field, count) in enumerate(top_fields):
        _write_tbl_row(fields_start + 2 + i, 2, 3,
                       field[:40], count, alt=bool(i % 2))

    # Bar chart — spans cols E-I (2nd+3rd box), vivid blue gradient
    if top_fields:
        from openpyxl.chart.series import DataPoint
        from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
        from copy import deepcopy

        chart2 = BarChart()
        chart2.type = "bar"
        chart2.title = "Top Research Fields"
        chart2.style = 10
        data_ref = Reference(ws, min_col=3, min_row=fields_start + 1,
                             max_row=fields_start + 1 + len(top_fields))
        cats_ref = Reference(ws, min_col=2, min_row=fields_start + 2,
                             max_row=fields_start + 1 + len(top_fields))
        chart2.add_data(data_ref, titles_from_data=True)
        chart2.set_categories(cats_ref)

        vivid_blues = [
            "0037B3", "0044CC", "0055E6", "0066FF", "1A75FF",
            "3385FF", "4D94FF", "66A3FF", "80B3FF", "99C2FF",
        ]
        series = chart2.series[0]
        for i in range(len(top_fields)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = vivid_blues[i % len(vivid_blues)]
            series.data_points.append(pt)

        chart2.width = 14
        chart2.height = 8
        chart2.legend = None
        ws.add_chart(chart2, f"E{fields_start + 1}")

    # ═══════════════════════════════════════════════════════════════════
    # TOP INSTITUTIONS TABLE (right side, cols K-L, row 28+)
    # ═══════════════════════════════════════════════════════════════════
    _write_merged(fields_start, 11, fields_start, 12, "Top Institutions",
                  font=section_font, fill=bg_fill, border=section_border)
    _write_tbl_header(fields_start + 1, 11, "Institution")
    _write_tbl_header(fields_start + 1, 12, "Jobs")

    for i, (inst, count) in enumerate(top_institutions):
        _write_tbl_row(fields_start + 2 + i, 11, 12,
                       inst[:35], count, alt=bool(i % 2))

    # ═══════════════════════════════════════════════════════════════════
    # FOOTER
    # ═══════════════════════════════════════════════════════════════════
    footer_row = (fields_start + 2
                  + max(len(top_fields), len(top_institutions), 1) + 2)
    ws.row_dimensions[footer_row].height = 18
    _write_merged(
        footer_row, 2, footer_row, 12,
        f"Generated by Job Search Pipeline   |   "
        f"Data covers {len(all_jobs)} positions across "
        f"{len(set(j.get('country') for j in all_jobs if j.get('country')))} "
        f"countries   |   {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        font=Font(size=9, color=TEXT_MUTED, name="Calibri"),
        fill=bg_fill,
        border=Border(top=Side(style="thin", color=SUBTLE_BORDER)))

    # Fill background for gutter columns
    for r in range(1, footer_row + 1):
        for c in (1, 4, 7, 10, 13):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                cell.fill = bg_fill


def _write_pi_headers_openpyxl(ws) -> None:
    """Write PI_COLUMNS as header row with formatting."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="16213e")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(vertical="center")

    pi_widths = {
        "PI Name": 20, "Institute": 25, "Country": 12, "Tier": 6,
        "h-index": 8, "Citations": 10, "Fields": 30, "Connected Seeds": 20,
        "Recommendation Score": 15, "Lab URL": 15, "Scholar URL": 15,
    }

    for col_idx, col_name in enumerate(PI_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = pi_widths.get(col_name, 15)

    ws.freeze_panes = "A2"


def _rebuild_pi_sheet_openpyxl(wb, rec_pis: list[dict]) -> None:
    """Delete and recreate the PI Recommendations sheet using openpyxl."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    if "PI Recommendations" in wb.sheetnames:
        wb.remove(wb["PI Recommendations"])

    ws = wb.create_sheet("PI Recommendations")
    _write_pi_headers_openpyxl(ws)

    link_font = Font(color="0563C1", underline="single")

    for row_idx, pi in enumerate(rec_pis, 2):
        row_data = _pi_to_row(pi)
        for col_idx, col_name in enumerate(PI_COLUMNS, 1):
            val = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_name in ("Lab URL", "Scholar URL"):
                if val and str(val).startswith("http"):
                    cell.hyperlink = str(val)
                    cell.font = link_font
            if col_name == "Tier" and val:
                _apply_tier_format_openpyxl(cell, val)

    # Auto-filter
    if ws.max_row and ws.max_row > 1:
        last_col = get_column_letter(len(PI_COLUMNS))
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


# ---------------------------------------------------------------------------
# Export paths: _fresh_export (xlsxwriter) and _incremental_update (openpyxl)
# ---------------------------------------------------------------------------

def _fresh_export(
    filepath: Path,
    us_jobs: list[dict], eu_jobs: list[dict],
    korea_jobs: list[dict], other_jobs: list[dict],
    all_jobs: list[dict], rec_pis: list[dict],
) -> Path:
    """Fresh export using xlsxwriter (full rewrite, no previous file)."""
    df_us = pd.DataFrame([_job_to_row(j) for j in us_jobs], columns=JOB_COLUMNS)
    df_eu = pd.DataFrame([_job_to_row(j) for j in eu_jobs], columns=JOB_COLUMNS)
    df_korea = pd.DataFrame([_job_to_row(j) for j in korea_jobs], columns=JOB_COLUMNS)
    df_other = pd.DataFrame([_job_to_row(j) for j in other_jobs], columns=JOB_COLUMNS)

    with pd.ExcelWriter(str(filepath), engine="xlsxwriter") as writer:
        _write_summary_dashboard(writer, all_jobs, rec_pis)

        for sheet_name, df, jobs in [
            ("US Positions", df_us, us_jobs),
            ("EU Positions", df_eu, eu_jobs),
            ("Korea Positions", df_korea, korea_jobs),
            ("Other Positions", df_other, other_jobs),
        ]:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            if len(df) > 0:
                _style_worksheet(writer, sheet_name, df)
                _style_deadline_cells(writer, sheet_name, df)
                _write_paper_links(writer, sheet_name, df, jobs)
                _style_citizenship_cells(writer, sheet_name, df, jobs)

        df_rec = pd.DataFrame(
            [_pi_to_row(p) for p in rec_pis], columns=PI_COLUMNS)
        df_rec.to_excel(writer, sheet_name="PI Recommendations", index=False)
        if len(df_rec) > 0:
            _style_worksheet(writer, "PI Recommendations", df_rec)

    return filepath


def _incremental_update(
    prev_path: Path, filepath: Path,
    new_jobs_by_region: dict, all_dismissed: set,
    all_jobs: list[dict], rec_pis: list[dict],
) -> Path:
    """Incremental update: load previous file with openpyxl, preserving formatting.

    Existing cell formatting (bold, background color, text color, user notes)
    is fully preserved.  Only dismissed rows are removed and new jobs appended.
    """
    from openpyxl import load_workbook

    wb = load_workbook(str(prev_path))

    sheet_region_map = {
        "US Positions": "US",
        "EU Positions": "EU",
        "Korea Positions": "Korea",
        "Other Positions": "Other",
    }

    total_new = 0
    for sheet_name, region_key in sheet_region_map.items():
        new_jobs = _update_sheet(wb, sheet_name, region_key,
                                 new_jobs_by_region, all_dismissed)
        total_new += len(new_jobs)

    logger.info("Incremental update: %d new jobs appended "
                "(formatting preserved)", total_new)

    # Rebuild auto-generated sheets (no user formatting to preserve)
    _rebuild_summary_dashboard_openpyxl(wb, all_jobs, rec_pis)
    _rebuild_pi_sheet_openpyxl(wb, rec_pis)

    wb.save(str(filepath))
    return filepath


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def export_to_excel(output_dir: Path = None, full_refresh: bool = False) -> Path:
    """Export job data to a multi-sheet Excel file.

    Split path:
    - full_refresh=True   → reset all statuses, clear exported_at, delete
                            existing Excel, export ALL active jobs via
                            _fresh_export() with hard exclusion only
    - Previous file exists → _incremental_update() (openpyxl, preserves formatting)
    - No previous file    → _fresh_export() (xlsxwriter, full generation)
    """
    output_dir = output_dir or EXCEL_OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    filepath = output_dir / "JobSearch_Auto.xlsx"

    if full_refresh:
        # ── Full refresh: reset DB, delete old file, export everything ──
        from src.db import get_connection
        with get_connection() as conn:
            conn.execute("UPDATE jobs SET status = 'new' WHERE status IN ('dismissed','merged')")
            conn.execute("UPDATE jobs SET exported_at = NULL")
        logger.info("Full refresh: reset all statuses and exported_at")

        # Delete existing file so we get a clean _fresh_export()
        if filepath.exists():
            filepath.unlink()
            logger.info("Full refresh: deleted existing %s", filepath)

        # Load all jobs, apply hard exclusion only (EXCLUDE_KEYWORDS in description)
        raw_jobs = get_jobs(limit=10000)
        all_jobs = [j for j in raw_jobs if not _is_excluded(j)]
        all_dismissed: set[str] = set()

        excluded_count = len(raw_jobs) - len(all_jobs)
        logger.info("Full refresh: %d total → %d after exclusion (%d excluded)",
                    len(raw_jobs), len(all_jobs), excluded_count)
    else:
        # ── Incremental (existing behaviour) ──
        # Phase 1: Read previous Excel (values only, for dismiss detection)
        prev_sheets = _read_previous_excel(output_dir)

        # Phase 2: Detect user deletions from previous Excel
        raw_pre = get_jobs(limit=10000)
        _skip_statuses = {"dismissed"}
        pre_urls = {j.get("url") for j in raw_pre
                    if j.get("url") and j.get("status") not in _skip_statuses}
        dismissed = _detect_dismissed_urls(prev_sheets, pre_urls)

        # Also include all permanently dismissed URLs
        from src.db import get_dismissed_urls
        all_dismissed = get_dismissed_urls() | dismissed

        # Phase 3: Load all non-dismissed jobs from DB
        raw_jobs = get_jobs(limit=10000)
        all_jobs = [j for j in raw_jobs
                    if j.get("status") not in _skip_statuses
                    and j.get("url") not in all_dismissed]

        # Apply exclusion filter (expired deadlines, past-year posts, keyword exclusions)
        pre_filter = len(all_jobs)
        excluded_urls = {j.get("url") for j in all_jobs
                         if _is_excluded(j) and j.get("url")}
        all_jobs = [j for j in all_jobs if not _is_excluded(j)]
        excluded_count = pre_filter - len(all_jobs)
        # Add excluded URLs to dismissed set so _update_sheet removes them
        # from existing Excel rows too (not just new jobs)
        all_dismissed = all_dismissed | excluded_urls
        logger.info("Loaded %d → %d jobs (%d excluded, %d dismissed)",
                    len(raw_jobs), len(all_jobs), excluded_count, len(all_dismissed))

    # Phase 4: Split by region and sort
    us_jobs = _sort_by_tier([j for j in all_jobs if j.get("region") == "US"])
    eu_jobs = _sort_by_tier([j for j in all_jobs if j.get("region") == "EU"])
    korea_jobs = _sort_by_tier([j for j in all_jobs if j.get("region") == "Korea"])
    other_jobs = _sort_by_tier([j for j in all_jobs if j.get("region") not in ("US", "EU", "Korea")])

    rec_pis = _sort_pis_by_tier(get_recommended_pis())

    # Phase 5: Choose export path
    if full_refresh:
        # Always do a fresh export on full refresh
        filepath = _fresh_export(filepath, us_jobs, eu_jobs, korea_jobs,
                                 other_jobs, all_jobs, rec_pis)
    else:
        prev_path = _find_previous_excel(output_dir)
        if prev_path:
            new_jobs_by_region = {
                "US": us_jobs, "EU": eu_jobs,
                "Korea": korea_jobs, "Other": other_jobs,
            }
            filepath = _incremental_update(prev_path, filepath, new_jobs_by_region,
                                           all_dismissed, all_jobs, rec_pis)
        else:
            filepath = _fresh_export(filepath, us_jobs, eu_jobs, korea_jobs,
                                     other_jobs, all_jobs, rec_pis)

    # Mark all exported jobs so we can detect user deletions next time
    exported_urls = [j.get("url") for j in all_jobs if j.get("url")]
    _mark_exported(exported_urls)

    total_rows = len(us_jobs) + len(eu_jobs) + len(korea_jobs) + len(other_jobs)
    logger.info("Excel exported to %s (%d total rows)", filepath, total_rows)
    return filepath
