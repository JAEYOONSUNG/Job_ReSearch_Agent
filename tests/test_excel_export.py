"""Tests for src.reporting.excel_export."""

from datetime import date
from unittest.mock import patch

from openpyxl import load_workbook

from src.db import dismiss_urls, get_connection, upsert_job
from src.reporting.excel_export import _is_excluded, _job_to_row, export_to_excel


def _sheet_urls(workbook_path, sheet_name: str) -> list[str]:
    """Return all non-empty Job URL values from a worksheet."""
    wb = load_workbook(workbook_path, read_only=True)
    try:
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        url_idx = headers.index("Job URL")
        return [
            row[url_idx]
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[url_idx]
        ]
    finally:
        wb.close()


def test_full_refresh_excludes_dismissed_jobs(
    test_db, tmp_path, sample_job_synbio, sample_job_crispr
):
    """Full refresh should preserve and exclude user-dismissed jobs."""
    kept_job = {**sample_job_synbio, "status": "new"}
    dismissed_job = {**sample_job_crispr, "status": "new"}

    upsert_job(kept_job)
    upsert_job(dismissed_job)
    dismiss_urls([dismissed_job["url"]])

    with patch("src.reporting.excel_export._is_excluded", return_value=False):
        workbook_path = export_to_excel(output_dir=tmp_path, full_refresh=True)

    exported_urls = _sheet_urls(workbook_path, "US Positions")
    assert kept_job["url"] in exported_urls
    assert dismissed_job["url"] not in exported_urls

    with get_connection() as conn:
        status = conn.execute(
            "SELECT status FROM jobs WHERE url = ?",
            (dismissed_job["url"],),
        ).fetchone()["status"]
        assert status == "dismissed"


def test_korea_export_wraps_long_cells_and_fills_structured_columns(test_db, tmp_path):
    """Korean rows should populate structured fields and preserve wrapped display."""
    korean_job = {
        "title": "KRIBB 박사후연구원 모집",
        "pi_name": "김철수",
        "institute": "KRIBB",
        "department": "합성생물학연구센터",
        "country": "South Korea",
        "region": "Korea",
        "field": "Synthetic Biology",
        "description": (
            "모집 내용:\n"
            "Bridge RNA 기반 유전자 조절 플랫폼 개발\n\n"
            "담당 업무:\n"
            "- 세포주 제작\n"
            "- CRISPR functional assay 구축\n\n"
            "지원 자격:\n"
            "- 생명과학 관련 박사학위\n"
            "- 분자생물학 경험\n\n"
            "우대 사항:\n"
            "- NGS 및 bioinformatics 경험\n\n"
            "근무 조건:\n"
            "- 연봉 6,000만원\n"
            "- 계약기간 2년\n\n"
            "제출 서류:\n"
            "- 이력서\n"
            "- 자기소개서\n"
            "- 연구계획서\n"
            "- 추천서\n"
        ),
        "url": "https://example.com/jobs/korea-bridge-rna",
        "source": "korean_jobs:hibrain",
        "status": "new",
    }
    upsert_job(korean_job)

    with patch("src.reporting.excel_export._is_excluded", return_value=False):
        workbook_path = export_to_excel(output_dir=tmp_path, full_refresh=True)

    wb = load_workbook(workbook_path)
    try:
        ws = wb["Korea Positions"]
        headers = [cell.value for cell in ws[1]]
        row = {headers[i]: ws.cell(row=2, column=i + 1).value for i in range(len(headers))}

        assert "Bridge RNA" in row["Position Summary"]
        assert "CRISPR functional assay" in row["Responsibilities"]
        assert "박사학위" in row["Requirements (Full)"]
        assert "bioinformatics" in row["Preferred Qualifications"]
        assert "CV" in row["Application Materials"]

        desc_col = headers.index("Description") + 1
        assert ws.cell(row=2, column=desc_col).alignment.wrapText is True
        assert (ws.row_dimensions[2].height or 0) > 20
    finally:
        wb.close()


def test_job_to_row_classifies_structured_sections_without_multiline_spill():
    """Structured fields should be flattened into their target columns."""
    row = _job_to_row({
        "title": "Postdoctoral Researcher",
        "country": "South Korea",
        "region": "Korea",
        "description": (
            "모집 내용:\n"
            "Bridge RNA 기반 유전자 조절 기술 개발\n\n"
            "담당 업무:\n"
            "- CRISPR assay development\n"
            "- NGS data analysis\n\n"
            "지원 자격:\n"
            "- 생명과학 관련 박사학위\n"
            "- 분자생물학 경험\n\n"
            "우대 사항:\n"
            "- Python 경험\n\n"
            "근무 조건:\n"
            "- 연봉 6,000만원\n"
            "- 계약기간 2년\n"
            "- 정규직\n"
            "- 채용 즉시\n\n"
            "제출 서류:\n"
            "- 이력서\n"
            "- 자기소개서\n"
            "- 추천서\n"
        ),
        "url": "https://example.com/jobs/structured-1",
    })

    assert row["Position Summary"] == "Bridge RNA 기반 유전자 조절 기술 개발"
    assert "\n" not in row["Responsibilities"]
    assert "CRISPR assay development" in row["Responsibilities"]
    assert row["Preferred Qualifications"] == "Python 경험"
    assert "우대 사항" not in row["Requirements (Full)"]
    assert "박사학위" in row["Degree Required"]
    assert row["Salary"] == "연봉 6,000만원"
    assert row["Duration"] == "2년"
    assert row["Contract Type"] == "Permanent"
    assert row["Start Date"] == "채용 즉시"
    assert "CV" in row["Application Materials"]


def test_job_to_row_falls_back_to_title_for_summary_and_field():
    row = _job_to_row({
        "title": "Postdoctoral Researcher in Neuroscience",
        "country": "South Korea",
        "region": "Korea",
        "url": "https://example.com/jobs/title-only",
    })

    assert row["Position Summary"] == "Postdoctoral Researcher in Neuroscience"
    assert row["Field"] == "Neuroscience"


def test_job_to_row_infers_doctorate_from_postdoc_title():
    row = _job_to_row({
        "title": "UNIST 생명과학과 Post-Doc. 모집",
        "country": "South Korea",
        "region": "Korea",
        "url": "https://example.com/jobs/hibrain-postdoc",
    })

    assert row["Degree Required"] == "Doctorate"


def test_is_excluded_drops_old_year_titles():
    past_year = date.today().year - 1
    assert _is_excluded({"title": f"{past_year}년 제1차 박사후연구원 채용공고"})
